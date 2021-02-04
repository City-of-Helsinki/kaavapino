import logging
import re
from typing import Iterable, Sequence, List

from openpyxl import load_workbook
from django.db import transaction

from users.models import PRIVILEGE_LEVELS
from projects.models import (
    Attribute,
    Deadline,
    DateType,
    AutomaticDate,
    DateCalculation,
    DeadlineDateCalculation,
    DeadlineDistance,
    ProjectSubtype,
    ProjectPhase,
)

logger = logging.getLogger(__name__)

# Sheets
DEADLINES_SHEET_NAME = "Aikatauluetapit"
DATETYPES_SHEET_NAME = "Päivätyypit"

DEADLINES_A1_EXPECTED = "projektitietotunniste"
DEADLINE_CREATED_AT_ATTRIBUTE_FIELD_VALUE = "projektin_kaynnistys_pvm"

# Deadline sheet column titles
DEADLINE_ATTRIBUTE_IDENTIFIER = "projektitietotunniste"
DEADLINE_ABBREVIATION = "etapin lyhenne"
DEADLINE_EDIT_PRIVILEGE = "kuka/mikä muokkaa tietoa"
DEADLINE_UPDATE_CALCULATIONS = "mihin tietoon tieto kytkeytyy"
DEADLINE_MINIMUM_DISTANCE = "minimietäisyys"
DEADLINE_INITIAL_CALCULATIONS = "generoitu ehdotus"
DEADLINE_ATTRIBUTE_CONDITION = "milloin tieto kuuluu prosessin"
DEADLINE_DATE_TYPE = "tiedon päivätyyppi"
DEADLINE_CALCULATION_DATE_TYPE = "laskettavien päivien päivätyyppi"
DEADLINE_TYPE = "tiedon jananäkymätyyppi "
DEADLINE_PHASE = "vaihe, johon päivämäärä liittyy"
DEADLINE_ERROR_PAST_DUE = "mitä tapahtuu, jos aikatauluun merkittyä  päivämäärää ei ole vahvistettu ja kyseisen etapin määräaika on ohitettu"
DEADLINE_ERROR_MIN_DISTANCE_PREV = "virheilmoitus, jos minimietäisyys edelliseen etappiin ei täyty, kun käyttäjä editoi aikataulua "
DEADLINE_WARNING_MIN_DISTANCE_NEXT = "virheilmoitus, jos minimietäisyys seuraavaan etappiin ei täyty , kun käyttäjä editoi aikataulua "


# Date type row indices
DATETYPE_NAME_INDEX = 0
DATETYPE_EXCLUDE_TYPE_INDEX = 1
DATETYPE_BASE_DATETYPE_INDEX = 2

# Mappings
DEADLINE_TYPES = {
    "vaiheen alkupiste": Deadline.TYPE_PHASE_START,
    "vaiheen päätepiste": Deadline.TYPE_PHASE_END,
    "katkoviivan alkupiste": Deadline.TYPE_DASHED_START,
    "katkoviivan päätepiste": Deadline.TYPE_DASHED_END,
    "sisäpalkin alkupiste": Deadline.TYPE_INNER_START,
    "sisäpalkin päätepiste": Deadline.TYPE_INNER_END,
    "määräaikaetappi": Deadline.TYPE_MILESTONE,
}


class DeadlineImporterException(Exception):
    pass


class DeadlineImporter:
    """Import deadlines and date types from an Excel file"""
    def __init__(self, options=None):
        self.options = options
        self.workbook = None

    def _open_workbook(self, filename):
        try:
            return load_workbook(filename, read_only=True, data_only=True)
        except FileNotFoundError as e:
            raise DeadlineImporterException(e)

    def _extract_data_from_workbook(self, workbook):
        try:
            deadlines_sheet = workbook[DEADLINES_SHEET_NAME]
            datetypes_sheet = workbook[DATETYPES_SHEET_NAME]
        except KeyError as e:
            raise DeadlineImporterException(e)

        deadlines_a1_value = deadlines_sheet["A1"].value
        invalid_sheet_exception = DeadlineImporterException(
            "This does not seem to be a valid attribute sheet."
        )
        try:
            if deadlines_a1_value.lower() != DEADLINES_A1_EXPECTED:
                raise invalid_sheet_exception
        except AttributeError:
            raise invalid_sheet_exception

        return (
            self._rows_for_deadline_sheet(deadlines_sheet),
            self._rows_for_datetype_sheet(datetypes_sheet),
        )

    def _rows_for_deadline_sheet(self, sheet):
        data = []

        for row in list(sheet.iter_rows()):
            if not row[1].value:
                break
            data.append([col.value for col in row])

        return data

    def _rows_for_datetype_sheet(self, sheet):
        data = []

        last_row = int(re.findall(r"[0-9]+", sheet.calculate_dimension())[-1])

        for i, row in enumerate(list(sheet.iter_rows())):
            if i > last_row:
                break
            data.append([col.value for col in row])

        return data

    def _set_row_indexes(self, header_row: Iterable[str]):
        """Determine index number for all columns."""
        self.column_index: dict = {}

        for index, column in enumerate(header_row):
            if column:
                self.column_index[column.lower()] = index

    def _check_if_row_valid(self, row: Sequence) -> bool:
        """Check if the row has all required data."""
        try:
            assert(row[self.column_index[DEADLINE_ABBREVIATION]])
            assert(row[self.column_index[DEADLINE_PHASE]])
        except AssertionError:
            abbreviation = row[self.column_index[DEADLINE_ABBREVIATION]]
            logger.info(f"Invalid deadline {abbreviation} not imported.")
            return False

        return True

    # Return a list of tuples with conditions and branches
    def _parse_conditions(self, rule):
        strings = list(zip(
            re.findall(r"\{%\s*if\s*(.*?)\s*%\}.*?\{%\s*endif\s*%\}", rule),
            re.findall(r"\{%+\sif.*?%\}\s*(.*?)\s*\{% endif %\}", rule),
        ))

        return [
            (re.split(r"\sor\s", cond), branch)
            for cond, branch in strings
        ]

    def _create_business_days_datetype(self):
        DateType.objects.update_or_create(
            name="Arkipäivät",
            identifier="arkipäivät",
            defaults={
                "exclude_selected": True,
                "business_days_only": True,
            },
        )

    def _create_datetypes(self, datetype_rows):
        def create_automatic_date(date_string):
            field_values = {}

            # Get weekdays
            try:
                weekdays = [
                    ["ma", "ti", "ke", "to", "pe", "la", "su"].index(weekday)
                    for weekday in re.findall(
                        r"(ma|ti|ke|to|pe|la|su)",
                        re.findall(r"\((.*)\)", date_string)[0],
                    )
                ]
            except IndexError:
                weekdays = None

            if weekdays:
                field_values["weekdays"] = weekdays
            else:
                field_values["weekdays"] = [0,1,2,3,4,5,6]


            # Get rest of the rule
            try:
                week = re.findall(r"vko ([0-9]*)", date_string)[0]
            except IndexError:
                week = None

            dates = re.findall(r"([0-9]+\.[0-9]*)", date_string)

            # AutomaticDate allows for more complexity, but for now it's enough
            # to always import single dates as one-day date ranges
            start_date = None
            end_date = None

            if dates:
                start_date = dates[0]
                try:
                    end_date = dates[1]
                except IndexError:
                    end_date = start_date

                start_month = start_date.split(".")[1]
                end_month = end_date.split(".")[1]

                if start_month and end_month:
                    pass
                elif not start_month and end_month:
                    month = end_date.split(".")[1]
                    if end_date and month:
                        start_date += month
                else:
                    logger.warning(
                        f"Missing month in date range {date_string}, ignoring."
                    )
                    return None

            after_holiday = None
            before_holiday = None

            # Handle special cases
            if date_string == "Pääsiäisen jälkeinen tiistai":
                weekdays = [1]
                after_holiday = "Easter Sunday"
            elif date_string == "Juhannusta edeltävä tiistai ja kesäkuu sen jälkeen":
                # In other words "Every Tuesday between June 16th and 30th"
                weekdays = [1]
                start_date = "16.6."
                end_date = "30.6."

            # Create AutomaticDate
            if week:
                field_values["week"] = week
            elif before_holiday:
                field_values["before_holiday"] = before_holiday
            elif after_holiday:
                field_values["after_holiday"] = after_holiday
            elif start_date or end_date:
                if start_date:
                    field_values["start_date"] = start_date
                if end_date:
                    field_values["end_date"] = end_date
            else:
                field_values["start_date"] = "1.1."
                field_values["end_date"] = "31.12."

            return AutomaticDate.objects.create(
                name=date_string,
                **field_values,
            )

        logger.info("Creating date types")

        names = datetype_rows[DATETYPE_NAME_INDEX][2:]
        identifiers = [
            value.lower().replace(" ", "_")
            for value in datetype_rows[DATETYPE_NAME_INDEX][2:]
        ]
        exclude_selecteds = [
            value == "kaikki paitsi"
            for value in datetype_rows[DATETYPE_EXCLUDE_TYPE_INDEX][2:]
        ]
        base_datetypes = [
            value.lower().replace(" ", "_") if value else None
            for value in datetype_rows[DATETYPE_BASE_DATETYPE_INDEX][2:]
        ]
        date_lists = [[] for _ in names]

        for i in list(range(len(names))):
            col_index = i + 2
            for row in datetype_rows[4:]:
                if row[col_index]:
                    date_lists[i].append(row[col_index])
                else:
                    break

        datetypes = zip(
            names,
            identifiers,
            exclude_selecteds,
            base_datetypes,
            date_lists,
        )

        for (name, identifier, exclude_selected, base_datetype, date_list) \
            in datetypes:

            datetype, _ = DateType.objects.update_or_create(
                identifier=identifier,
                defaults={
                    "name": name,
                    "exclude_selected": exclude_selected,
                },
            )

            try:
                datetype.base_datetype.set([DateType.objects.get(
                    identifier=base_datetype,
                )])
            except DateType.DoesNotExist:
                logger.warning(
                    f"Ignored invalid base date type {base_datetype} for date type {name}."
                )

            datetype.automatic_dates.all().delete()

            automatic_dates = [
                create_automatic_date(date)
                for date in date_list
            ]
            automatic_dates = [date for date in automatic_dates if date]

            datetype.automatic_dates.set(automatic_dates)

    def get_datetype(self, row, target):
        if row:
            try:
                identifier = row.lower().replace(" ", "_")
                return DateType.objects.get(
                    identifier=identifier
                )
            except DateType.DoesNotExist:
                logger.warning(
                    f"Ignoring invalid date type {row} for {target}."
                )

        return None

    def _create_deadlines(self, subtype, rows):
        logger.info(f"Creating deadlines for {subtype}")

        for i, row in enumerate(rows):
            abbreviation = row[self.column_index[DEADLINE_ABBREVIATION]]
            attribute = row[self.column_index[DEADLINE_ATTRIBUTE_IDENTIFIER]]

            if attribute == DEADLINE_CREATED_AT_ATTRIBUTE_FIELD_VALUE:
                default_to_created_at = True
            else:
                default_to_created_at = False

            try:
                attribute = Attribute.objects.get(
                    identifier=attribute
                )
            except Attribute.DoesNotExist:
                logger.warning(
                    f"Ignored invalid attribute identifier {attribute} for deadline {abbreviation}."
                )
                attribute = None

            deadline_types = []
            for dl_type in re.split(
                r";[\s]*", row[self.column_index[DEADLINE_TYPE]] or ""
            ):
                try:
                    deadline_types.append(DEADLINE_TYPES[dl_type])
                except KeyError:
                    pass

            date_type = self.get_datetype(
                row[self.column_index[DEADLINE_DATE_TYPE]],
                f"deadline {abbreviation}",
            )

            condition_attributes = []

            # check possible subtype limitations
            split_cell = re.split(
                r";\s*", row[self.column_index[DEADLINE_ATTRIBUTE_CONDITION]]
            )
            subtypes = re.findall(
                r"XS|S|M|L|XL",
                re.split(r"\s*==\s*|\s+in\s+", split_cell[0])[-1],
            )
            if len(split_cell) < 2 or subtype.name in subtypes:
                cond_attr_identifiers = self._parse_conditions(
                    row[self.column_index[DEADLINE_ATTRIBUTE_CONDITION]]
                )

            if not len(cond_attr_identifiers) \
                and len(subtypes) \
                and subtype.name not in subtypes:
                continue

            for identifier in cond_attr_identifiers:
                try:
                    condition_attributes.append(
                        Attribute.objects.get(identifier=identifier)
                    )
                except Attribute.DoesNotExist:
                    logger.warning(
                        f"Ignored invalid condition attribute {identifier} for deadline {abbreviation}."
                    )

            try:
                phase = subtype.phases.get(name=row[self.column_index[DEADLINE_PHASE]])
            except ProjectPhase.DoesNotExist:
                logger.warning(
                    f"Invalid phase {row[self.column_index[DEADLINE_PHASE]]} for deadline {abbreviation} in {subtype}, skipping."
                )
                continue

            error_past_due = row[self.column_index[DEADLINE_ERROR_PAST_DUE]]
            error_min_distance_previous = row[self.column_index[DEADLINE_ERROR_MIN_DISTANCE_PREV]]
            warning_min_distance_next = row[self.column_index[DEADLINE_WARNING_MIN_DISTANCE_NEXT]]
            index = i + 1

            deadline, _ = Deadline.objects.update_or_create(
                abbreviation=abbreviation,
                phase=phase,
                subtype=subtype,
                defaults={
                    "attribute": attribute,
                    "deadline_types": deadline_types,
                    "date_type": date_type,
                    "error_past_due": error_past_due,
                    "error_min_distance_previous": error_min_distance_previous,
                    "warning_min_distance_next": warning_min_distance_next,
                    "default_to_created_at": default_to_created_at,
                    "index": index,
                },
            )

            deadline.condition_attributes.set(condition_attributes)

    def _create_deadline_relations(self, subtype, rows):
        logger.info(f"Updating deadline relations for {subtype}")

        def parse_and_create_calculations(calc_string, calc_datetype):
            abbreviation_regex = r"([A-Z]+[0-9]+\.?[0-9]*)\s*[+|-]*\s*[0-9]*"
            constant_regex = r"[+|-]\s*([0-9]*)"
            identifier_regex = r"\{\{(.*)\}\}"
            calculations = []

            conditions_parsed = self._parse_conditions(calc_string) or [ \
                ([], calc) for calc in re.split(r";\s*", calc_string)]

            for index, (conds, calc) in enumerate(conditions_parsed):
                condition_attributes = []
                not_condition_attributes = []
                subtype_conds = [
                    cond for cond in conds
                    if cond[:25] == "kaavaprosessin_kokoluokka"
                ]
                attribute_conds = [
                    cond for cond in conds
                    if cond[:25] != "kaavaprosessin_kokoluokka"
                ]

                # Handle subtype conditions first and discard branches
                # that do not match
                subtype_matches = not subtype_conds

                for cond in subtype_conds:
                    subtypes = re.findall(
                        r"XS|S|M|L|XL",
                        re.split(r"\s*==\s*|\s+in\s+", cond)[-1],
                    )
                    if subtype.name in subtypes:
                        subtype_matches = True
                        break

                if not subtype_matches:
                    continue

                # Other valid conditions are saved as Attribute relations later
                for cond in attribute_conds:
                    negate = False

                    if cond[0] == "!":
                        negate = True
                        cond = cond[1:]

                    try:
                        attribute = Attribute.objects.get(identifier=cond)

                        if negate:
                           not_condition_attributes.append(attribute)
                        else:
                           condition_attributes.append(attribute)

                    except Attribute.DoesNotExist:
                        logger.warning(
                            f"Ignored an invalid attribute identifier {cond} for calculating deadline {abbreviation}."
                        )

                try:
                    constant = int(re.findall(constant_regex, calc)[0])
                    if re.findall(r"[+|-]", calc)[0] == "-":
                        constant = -constant
                except IndexError:
                    constant = 0
                except ValueError:
                    logger.warning(
                        f"Ignored invalid calculation {calc} for deadline {abbreviation}."
                    )
                    continue

                try:
                    base_deadline = Deadline.objects.get(
                        abbreviation=re.findall(abbreviation_regex, calc)[0],
                        subtype=subtype,
                    )
                except Deadline.DoesNotExist:
                    logger.warning(
                        f"Ignored an invalid deadline abbreviation {re.findall(abbreviation_regex, calc)[0]} for calculating deadline {abbreviation}."
                    )
                    base_deadline = None
                except IndexError:
                    base_deadline = None

                try:
                    base_attribute = Attribute.objects.get(
                        identifier=re.findall(identifier_regex, calc)[0],
                    )
                except Attribute.DoesNotExist:
                    logger.warning(
                        f"Ignored an invalid attribute identifier {re.findall(identifier_regex, calc)[0]} for calculating deadline {abbreviation}."
                    )
                    base_attribute = None
                except IndexError:
                    base_attribute = None

                if not base_attribute and not base_deadline:
                    logger.warning(
                        f"Ignored invalid calculation {calc} for deadline {abbreviation}."
                    )
                    continue

                calc_object = DeadlineDateCalculation.objects.create(
                    deadline=deadline,
                    datecalculation=DateCalculation.objects.create(
                        base_date_attribute=base_attribute,
                        base_date_deadline=base_deadline,
                        constant=constant,
                        date_type=calc_datetype,
                    ),
                    index=index,
                )
                calc_object.conditions.set(condition_attributes)
                calc_object.not_conditions.set(not_condition_attributes)
                calculations.append(calc_object)

            return calculations

        """
        Supported formats for distance string:
        "xx"
        "xx + a"
        "xx - a"
        "xx - a; yy - b"
        "{% if kaavaprosessin_kokoluokka == S %} xx + a {% endif %}"
        "{% if kaavaprosessin_kokoluokka in [M, L, XL] %} xx + as {% endif %}"
        """
        def parse_distance(distance_string):
            def add_distance(calc, attr=None):
                try:
                    try:
                        [reference, operator, distance] = \
                            re.split(r"\s*([-, +])\s*", calc)
                    except ValueError:
                        [reference, operator, distance] = [calc, "+", 0]

                    distances.append((
                        Deadline.objects.get(subtype=subtype, abbreviation=reference),
                        operator,
                        distance,
                        [attr] if attr else [],
                    ))

                except Deadline.DoesNotExist:
                    logger.warning(
                        f"Ignored an invalid minimum distance reference {calc} for deadline {abbreviation}."
                    )

                return distances

            distances = []
            conditions_parsed = self._parse_conditions(distance_string)
            if not conditions_parsed:
                for calc in re.split(r";\s*", distance_string):
                    distances = add_distance(calc)

            else:
                for cond, calc in conditions_parsed:
                    if len(cond) < 1:
                        logger.warning(
                            f"Only one condition per distance rule currently supported in importer, ignored conditions {cond} for deadline {abbreviation}."
                        )
                        continue

                    cond = cond[0]
                    if cond and cond[:25] == "kaavaprosessin_kokoluokka":
                        subtypes = re.findall(
                            r"XS|S|M|L|XL",
                            re.split(r"\s*==\s*|\s+in\s+", cond)[-1],
                        )
                        if subtype.name in subtypes:
                            add_distance(calc)

                    elif not cond:
                        add_distance(calc)
                    else:
                        try:
                            attr = Attribute.objects.get(identifier=cond)
                            add_distance(calc, attr)
                        except (Attribute.DoesNotExist, AssertionError):
                            logger.warning(
                                f"Ignored an invalid minimum distance specification {cond} for deadline {abbreviation}; attribute not found."
                            )

            return distances

        for row in rows:
            try:
                abbreviation = row[self.column_index[DEADLINE_ABBREVIATION]]
                deadline = Deadline.objects.get(
                    subtype=subtype,
                    abbreviation=abbreviation,
                )
            except Deadline.DoesNotExist:
                continue

            # check possible subtype limitations
            split_cell = re.split(
                r";\s*", row[self.column_index[DEADLINE_ATTRIBUTE_CONDITION]]
            )
            subtypes = re.findall(
                r"XS|S|M|L|XL",
                re.split(r"\s*==\s*|\s+in\s+", split_cell[0])[-1],
            )
            if len(split_cell) < 2 or subtype.name in subtypes:
                cond_attr_identifiers = self._parse_conditions(
                    row[self.column_index[DEADLINE_ATTRIBUTE_CONDITION]]
                )

            if not len(cond_attr_identifiers) \
                and len(subtypes) \
                and subtype.name not in subtypes:
                continue

            calc_datetype = self.get_datetype(
                row[self.column_index[DEADLINE_CALCULATION_DATE_TYPE]],
                "deadline calculation",
            )

            # Create DateCalculations and Deadline relations
            if row[self.column_index[DEADLINE_INITIAL_CALCULATIONS]]:
                initial_calculations = parse_and_create_calculations(
                    row[self.column_index[DEADLINE_INITIAL_CALCULATIONS]],
                    calc_datetype,
                )
                deadline.initial_calculations.set(initial_calculations)

            if row[self.column_index[DEADLINE_UPDATE_CALCULATIONS]]:
                update_calculations = parse_and_create_calculations(
                    row[self.column_index[DEADLINE_UPDATE_CALCULATIONS]],
                    calc_datetype,
                )
                deadline.update_calculations.set(update_calculations)

            # Create minimum distance relations
            if row[self.column_index[DEADLINE_MINIMUM_DISTANCE]]:
                parsed = enumerate(parse_distance(
                    row[self.column_index[DEADLINE_MINIMUM_DISTANCE]]
                ))
                for index, (target, operator, distance, conditions) in parsed:
                    distance = DeadlineDistance.objects.create(
                        deadline=deadline if operator == "+" else target,
                        previous_deadline=target if operator == "+" else deadline,
                        distance_from_previous=distance,
                        index=index,
                    )
                    distance.conditions.set(conditions)
                    distance.save()
            else:
                logger.warning(
                    f"No minimum distance information found for {row[self.column_index[DEADLINE_ABBREVIATION]]}."
                )

    @transaction.atomic
    def run(self):
        if not Attribute.objects.count():
            raise DeadlineImporterException(
                "No Attributes found, run Attribute importer first"
            )
            return

        filename = self.options.get("filename")
        logger.info(f"Importing deadlines and datetypes from {filename}")

        self.workbook = self._open_workbook(filename)

        deadline_data_rows, datetype_data_rows = \
            self._extract_data_from_workbook(self.workbook)
        deadline_header_row = deadline_data_rows.pop(0)
        self._set_row_indexes(deadline_header_row)
        deadline_data_rows = [
            row for row in deadline_data_rows
            if self._check_if_row_valid(row)
        ]
        self._create_business_days_datetype()
        self._create_datetypes(datetype_data_rows)

        # Delete existing calculations and distances
        DeadlineDistance.objects.all().delete()
        DateCalculation.objects.all().delete()

        for subtype in ProjectSubtype.objects.all():
            self._create_deadlines(subtype, deadline_data_rows)
            self._create_deadline_relations(subtype, deadline_data_rows)
