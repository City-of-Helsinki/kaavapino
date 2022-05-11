import logging
import re
from typing import Iterable, Sequence

from openpyxl import load_workbook
from django.db import transaction

from projects.models import (
    ProjectType,
    ProjectSubtype,
    Report,
    ReportFilter,
    ReportFilterAttributeChoice,
    Attribute,
    Report,
    ReportColumn,
    ReportColumnPostfix,
)
from ..models.utils import create_identifier, truncate_identifier

logger = logging.getLogger(__name__)

IDENTIFIER_MAX_LENGTH = 50

# Sheets
REPORT_SHEET_NAME = "Sheet1"
FILTER_SHEET_NAME = "Sheet2"
EXPECTED_A1_VALUE = "rivi nro"

# Report sheet column titles
REPORT_NAME = "raportin nimi"
COLUMN_ATTRIBUTES = "kentät"
COLUMN_CONDITIONS = "näyttöehto"
COLUMN_POSTFIX = "loppuliite"
COLUMN_POSTFIX_ONLY = "piilota kentän arvo"
COLUMN_INDEX = "sarakkeiden järjestys"
COLUMN_TITLE = "sarakkeen otsikko"
COLUMN_PREVIEW = "näytetään esikatselussa"
COLUMN_PREVIEW_ONLY = "vain esikatselussa"
COLUMN_CUSTOM_VALUE_MAPPING = "korvaavat arvot vakioarvoille"
COLUMN_PREVIEW_TITLE_COLUMN = "saraketta käytetään esikatselun otsikkoina"
COLUMN_GENERATE_NEW_ROW = "luo uuden rivin"

# Filter sheet column titles
FILTER_NAME = "suodattimen nimi"
FILTER_REPORTS = "raportit joita koskee"
FILTER_ATTRIBUTES = "kenttätunnisteet"
FILTER_TYPE = "suodattimen tyyppi"
FILTER_INPUT_TYPE = "hakuarvon tyyppi"
FILTER_ATTRIBUTES_AS_CHOICES = "kenttätunnisteen valinta"
FILTER_ATTRIBUTE_CHOICE_VALUES = "kenttätunnisteen valinta, lisävalinta"

FILTER_TYPE_MAPPING = {
    "tarkka arvo": ReportFilter.TYPE_EXACT,
    "monivalinta": ReportFilter.TYPE_MULTIPLE,
    "arvoväli": ReportFilter.TYPE_RANGE,
    "arvo asetettu": ReportFilter.TYPE_SET,
    "arvoa ei asetettu": ReportFilter.TYPE_NOT_SET,
}

FILTER_INPUT_TYPE_MAPPING = {
    "teksti": ReportFilter.INPUT_TYPE_STRING,
    "oletus": ReportFilter.INPUT_TYPE_STRING,
    "henkilö (ad)": ReportFilter.INPUT_TYPE_PERSONNEL,
    "päivämäärä": ReportFilter.INPUT_TYPE_DATE,
    "kokonaisluku": ReportFilter.INPUT_TYPE_INTEGER,
}


class ReportImporterException(Exception):
    pass


class ReportImporter:
    """Import reports from an Excel file"""
    def __init__(self, options=None):
        self.options = options
        self.workbook = None

    def _open_workbook(self, filename):
        try:
            return load_workbook(filename, read_only=True, data_only=True)
        except FileNotFoundError as e:
            raise ReportImporterException(e)

    def _extract_data_from_workbook(self, workbook):
        try:
            report_sheet = workbook[REPORT_SHEET_NAME]
            filter_sheet = workbook[FILTER_SHEET_NAME]
        except KeyError as e:
            raise ReportImporterException(e)

        report_sheet_a1_value = report_sheet["A1"].value
        invalid_sheet_exception = ReportImporterException(
            "This does not seem to be a valid report type sheet."
        )
        try:
            if report_sheet_a1_value.lower() != EXPECTED_A1_VALUE:
                raise invalid_sheet_exception
        except AttributeError:
            raise invalid_sheet_exception

        return (
            self._rows_for_sheet(report_sheet),
            self._rows_for_sheet(filter_sheet),
        )

    def _rows_for_sheet(self, sheet):
        data = []

        for row in list(sheet.iter_rows()):
            if not row[0].value:
                break
            data.append([col.value for col in row])

        return data

    def _set_row_indexes(self, report_header_row, filter_header_row):
        """Determine index number for all columns."""
        self.report_column_index: dict = {}
        self.filter_column_index: dict = {}

        for index, column in enumerate(report_header_row):
            if column:
                self.report_column_index[column.lower()] = index

        for index, column in enumerate(filter_header_row):
            if column:
                self.filter_column_index[column.lower()] = index

    def _check_if_report_row_valid(self, row: Sequence) -> bool:
        """Check if the row has all required data."""
        try:
            assert(row[self.report_column_index[REPORT_NAME]])
        except AssertionError:
            logger.info(f"Invalid row missing report name not imported.")
            return False

        return True

    def _check_if_filter_row_valid(self, row: Sequence) -> bool:
        """Check if the row has all required data."""
        try:
            assert(row[self.filter_column_index[FILTER_NAME]])
            assert(row[self.filter_column_index[FILTER_TYPE]])
            assert(row[self.filter_column_index[FILTER_ATTRIBUTES]])
        except AssertionError:
            logger.info(f"Invalid row missing filter name, type, or attribute(s) not imported.")
            return False

        return True

    def _create_report_columns(self, rows):
        project_type, _ = ProjectType.objects.get_or_create(name="asemakaava")
        ReportColumn.objects.all().delete()
        ReportColumnPostfix.objects.all().delete()

        for row in rows:
            report_name = row[self.report_column_index[REPORT_NAME]]
            report, _ = Report.objects.get_or_create(
                name=report_name,
                project_type=project_type,
                defaults={
                    "is_admin_report": False,
                    "previewable": False,
                }
            )

            postfix_only = row[self.report_column_index[COLUMN_POSTFIX_ONLY]] == "kyllä"
            preview = row[self.report_column_index[COLUMN_PREVIEW]] == "kyllä"
            preview_only = row[self.report_column_index[COLUMN_PREVIEW_ONLY]] == "kyllä"
            if not report.previewable and (preview or preview_only):
                report.previewable = True
                report.save()

            mappings = row[self.report_column_index[COLUMN_CUSTOM_VALUE_MAPPING]]
            if mappings:
                custom_display_mapping = {
                    key: value for key, value in zip(
                        re.findall(r'"([^"]*)":', mappings),
                        re.findall(r':[\s]*"([^"]*)"', mappings),
                    )
                }
            else:
                custom_display_mapping = None

            column = ReportColumn.objects.create(
                report=report,
                postfix_only=postfix_only,
                preview=preview,
                preview_only=preview_only,
                title=row[self.report_column_index[COLUMN_TITLE]],
                index=row[self.report_column_index[COLUMN_INDEX]] or 0,
                preview_title_column=row \
                    [self.report_column_index[COLUMN_PREVIEW_TITLE_COLUMN]] == \
                    "kyllä",
                custom_display_mapping=custom_display_mapping,
                generates_new_rows=row \
                    [self.report_column_index[COLUMN_GENERATE_NEW_ROW]] == \
                    "kyllä",
            )

            attributes = \
                (row[self.report_column_index[COLUMN_ATTRIBUTES]] or "").split(",")
            conditions = \
                (row[self.report_column_index[COLUMN_CONDITIONS]] or "").split(",")

            column.attributes.set(
                Attribute.objects.filter(identifier__in=attributes)
            )
            column.condition.set(
                Attribute.objects.filter(identifier__in=conditions)
            )

            try:
                postfixes = zip(
                    re.findall(
                        r"\[([a-zA-Z_0-9!,]*)\]",
                        row[self.report_column_index[COLUMN_POSTFIX]],
                    ),
                    re.findall(
                        r'"([^"]*)"',
                        row[self.report_column_index[COLUMN_POSTFIX]],
                    ) or row[self.report_column_index[COLUMN_POSTFIX]],
                )
            except TypeError as e:
                postfixes = []

            for i, (rules, formatting) in enumerate(postfixes):
                rules = rules.split(",")
                subtype_options = ["XS", "S", "M", "L", "XL"]
                subtypes = []
                show_conditions = []
                show_not_conditions = []
                hide_conditions = []
                hide_not_conditions = []

                for rule in rules:
                    if rule in subtype_options:
                        subtypes.append(rule)
                    elif len(rule) >= 5 and rule[0:5] == "SHOW!":
                        show_not_conditions.append(rule[5:])
                    elif len(rule) >= 5 and rule[0:5] == "HIDE!":
                        hide_not_conditions.append(rule[5:])
                    elif len(rule) >= 4 and rule[0:4] == "SHOW":
                        show_conditions.append(rule[4:])
                    elif len(rule) >= 4 and rule[0:4] == "HIDE":
                        hide_conditions.append(rule[4:])
                    elif len(rule) and rule[0] == "!":
                        show_not_conditions.append(rule[1:])
                    else:
                        show_conditions.append(rule)

                if not subtypes:
                    subtypes = subtype_options

                postfix = ReportColumnPostfix.objects.create(
                    report_column=column,
                    formatting=formatting,
                    index = i,
                )
                postfix.subtypes.set(
                    ProjectSubtype.objects.filter(
                        name__in=subtypes,
                    )
                )
                postfix.show_conditions.set(
                    Attribute.objects.filter(
                        identifier__in=show_conditions,
                    )
                )
                postfix.hide_conditions.set(
                    Attribute.objects.filter(
                        identifier__in=hide_conditions,
                    )
                )
                postfix.show_not_conditions.set(
                    Attribute.objects.filter(
                        identifier__in=show_not_conditions,
                    )
                )
                postfix.hide_not_conditions.set(
                    Attribute.objects.filter(
                        identifier__in=hide_not_conditions,
                    )
                )

    def _clean_value_list(self, value):
        try:
            return value[0].strip(" ")
        except IndexError:
            return None

    def _create_report_filters(self, rows):
        ReportFilterAttributeChoice.objects.all().delete()
        ReportFilter.objects.all().delete()

        for row in rows:
            report_names = re.findall(
                r'"([^"]*)"',
                row[self.filter_column_index[FILTER_REPORTS]] or "",
            )
            reports = Report.objects.filter(name__in=report_names)

            attribute_identifiers = \
                row[self.filter_column_index[FILTER_ATTRIBUTES]].split(",")
            attributes = \
                Attribute.objects.filter(identifier__in=attribute_identifiers)

            name = row[self.filter_column_index[FILTER_NAME]]
            identifier = truncate_identifier(
                create_identifier(name.strip(" \t:.")),
                length=IDENTIFIER_MAX_LENGTH,
            )
            filter_type = \
                FILTER_TYPE_MAPPING[row[self.filter_column_index[FILTER_TYPE]]]
            filter_input_type = \
                FILTER_INPUT_TYPE_MAPPING[row[self.filter_column_index[FILTER_INPUT_TYPE]]]

            if row[self.filter_column_index[FILTER_ATTRIBUTE_CHOICE_VALUES]]:
                attributes_as_choices = True
                attr_vc_string = \
                    row[self.filter_column_index[FILTER_ATTRIBUTE_CHOICE_VALUES]]
                # import format:
                # ("attribute_identifier", "label head"): [("label tail", value, value), ...]
                attribute_choice_values = [
                    {
                        "identifier": re.findall(r'"([^"]*)"', key)[0],
                        "label_head": re.findall(r'"([^"]*)"', key)[1],
                        "value_choices": [
                            {
                                "label_tail": (re.findall(r'"([^"]*)"', val) + [""])[0],
                                "value": self._clean_value_list(
                                    re.findall(r'"[^"]*",\s*(.*)', val)
                                )
                            }
                            for val in re.findall(r'\(([^\(^\)]*)\)', values)
                        ],
                    }
                    for key, values in zip(
                        re.findall(r'\(([^\(^\)]*)\):', attr_vc_string),
                        re.findall(r'\([^\(^\)]*\):\s*\[([^\[^\]]*)\]', attr_vc_string),
                    )
                    if len(re.findall(r'"([^"]*)"', key)) == 2 and values
                ]

            elif row[self.filter_column_index[FILTER_ATTRIBUTES_AS_CHOICES]] == "kyllä":
                attributes_as_choices = True
                attribute_choice_values = []
            else:
                attributes_as_choices = False
                attribute_choice_values = []

            report_filter = ReportFilter.objects.create(
                name=name,
                identifier=identifier,
                type=filter_type,
                input_type=filter_input_type,
                attributes_as_choices=attributes_as_choices,
            )

            report_filter.reports.set(reports)
            report_filter.attributes.set(attributes)


            for attribute_choice in attribute_choice_values:
                for i, value_choice in enumerate(attribute_choice["value_choices"]):
                    if not value_choice["value"]:
                        continue

                    try:
                        ReportFilterAttributeChoice.objects.create(
                            report_filter=report_filter,
                            attribute=Attribute.objects.get(
                                identifier=attribute_choice["identifier"],
                            ),
                            name=f'{attribute_choice["label_head"]} {value_choice["label_tail"]}',
                            identifier=f'{attribute_choice["identifier"]}_{i}',
                            value=value_choice["value"],
                        )
                    except Attribute.DoesNotExist:
                        continue

    @transaction.atomic
    def run(self):
        filename = self.options.get("filename")
        logger.info(f"Importing report types and columns from {filename}")

        self.workbook = self._open_workbook(filename)

        report_data_rows, filter_data_rows = \
            self._extract_data_from_workbook(self.workbook)
        report_header_row = report_data_rows.pop(0)
        filter_header_row = filter_data_rows.pop(0)
        # the second row is reserved for column descriptions, remove it as well
        report_data_rows.pop(0)
        filter_data_rows.pop(0)
        self._set_row_indexes(report_header_row, filter_header_row)
        report_data_rows = [
            row for row in report_data_rows
            if self._check_if_report_row_valid(row)
        ]
        filter_data_rows = [
            row for row in filter_data_rows
            if self._check_if_filter_row_valid(row)
        ]
        Report.objects.all().delete()
        self._create_report_columns(report_data_rows)
        self._create_report_filters(filter_data_rows)
