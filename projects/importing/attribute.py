import itertools
import csv
import json
import logging
import re
from collections import Counter, defaultdict
from enum import Enum
from io import StringIO
from itertools import filterfalse
from collections.abc import Sequence
from typing import Iterable, List, Optional

from django.core.cache import cache
from django.db.utils import IntegrityError
from django.contrib.auth.models import Group
from django.db import transaction
from django.db.models import ProtectedError, Q
from openpyxl import load_workbook

from ..models import (
    Attribute,
    AttributeAutoValue,
    AttributeAutoValueMapping,
    AttributeValueChoice,
    DataRetentionPlan,
    Deadline,
    DocumentLinkFieldSet,
    DocumentLinkSection,
    FieldSetAttribute,
    OverviewFilter,
    OverviewFilterAttribute,
    ProjectCardSection,
    ProjectCardSectionAttribute,
    ProjectFloorAreaSection,
    ProjectFloorAreaSectionAttribute,
    ProjectFloorAreaSectionAttributeMatrixCell,
    ProjectFloorAreaSectionAttributeMatrixStructure,
    ProjectPhaseSection,
    ProjectPhaseSectionAttribute,
    ProjectPhase,
    CommonProjectPhase,
    Project,
    ProjectPhaseFieldSetAttributeIndex,
    ProjectType,
    ProjectSubtype,
    ProjectPhaseDeadlineSection,
    ProjectPhaseDeadlineSectionAttribute,
)
from ..models.attribute import AttributeCategorization
from ..models.utils import create_identifier, truncate_identifier, check_identifier

logger = logging.getLogger(__name__)

IDENTIFIER_MAX_LENGTH = 50
VALID_ATTRIBUTE_CALCULATION_TYPES = [Attribute.TYPE_DECIMAL, Attribute.TYPE_INTEGER]

PROJECT_SIZE = "prosessin kokoluokka, joissa kenttä näkyy"

DEFAULT_SHEET_NAME = "Kaavaprojektitiedot"
CHOICES_SHEET_NAME = "Pudotusvalikot"

ATTRIBUTE_NAME = "projektitieto"
ATTRIBUTE_IDENTIFIER = "projektitieto tunniste"
ATTRIBUTE_TYPE = "kenttätyyppi"
ATTRIBUTE_CHOICES_REF = "pudotusvalikko/vaihtoehdot"
ATTRIBUTE_UNIT = "mittayksikkö"
ATTRIBUTE_BROADCAST_CHANGES = "muutos näkyy viestit-ikkunassa"
ATTRIBUTE_REQUIRED = "pakollinen tieto"  # kyllä/ei
ATTRIBUTE_DATA_RETENTION = "tiedon säilytysaika"
ATTRIBUTE_MULTIPLE_CHOICE = "tietoa voi olla useita" # kyllä/ei
ATTRIBUTE_SEARCHABLE = "tietoa käytetään hakutietona (hakuruudussa) projektit-näkymässä"
ATTRIBUTE_RELATED_FIELDS = "mihin tietoon kytkeytyy"
ATTRIBUTE_LINKED_FIELDS = "mihin tieto siirtyy"
ATTRIBUTE_RULE_CONDITIONAL_VISIBILITY = "sääntö: kenttä näkyy vain toiseen kenttään tehdyn valinnan perusteella"
ATTRIBUTE_RULE_AUTOFILL = "sääntö: tieto muodostuu toiseen kenttään merkityn tiedon perusteella automaattisesti"
ATTRIBUTE_RULE_AUTOFILL_READONLY = "sääntö: voiko automaattisesti muodostunutta tietoa muokata "
ATTRIBUTE_RULE_UPDATE_AUTOFILL = "sääntö: vaikuttaako tiedon muokkaus aiemmin täytettyyn tietokenttään"
ATTRIBUTE_CHARACTER_LIMIT = "merkkien enimmäismäärä"
ATTRIBUTE_VALIDATION_REGEX = "regex sääntö"
ATTRIBUTE_HIGHLIGHT_GROUP = "korostettavat kentät"
ATTRIBUTE_EDIT_PRIVILEGE = "kenellä on oikeus muokata tietoa"
ATTRIBUTE_ERROR = "virhetilanne"
ATTRIBUTE_PLACEHOLDER = "syöttökentässä näkyvä ohjeistusteksti"
ATTRIBUTE_ASSISTIVE_TEXT = "kentän alla näkyvä lyhyt vinkki"
ATTRIBUTE_ERROR_TEXT = "virheteksti"
ATTRIBUTE_FIELD_ROLE = "rooli kenen kenttä"
ATTRIBUTE_FIELD_SUBROLE = "alirooli kenen kenttä"
# TODO: ask for a dedicated column for uniqueness at some point
ATTRIBUTE_ERROR_UNIQUE = [
    "Virhe. Nimi on jo käytössä",
    "Virhe. Diaarinumero on jo toisen projektin käytössä. Samalle diaarínumerolle ei voi luoda uutta projektia.",
]
ATTRIBUTE_AUTO_VALUE_KEY_FIELD = "yhteystietojen automaattisen täytön lähdekenttä"
ATTRIBUTE_AUTO_VALUE_MAPPING = "yhteystietojen automaattisen täytön avain-arvoparit"

# Attribute object mappings for static Project fields
STATIC_ATTRIBUTES_MAPPING = {
     "vastuuhenkilo_nimi": "user",
     "luodaanko_nakyvaksi": "public",
     "pinonumero": "pino_number",
     "projektin_nimi": "name",
     "periaatteet_luotu": "create_principles",
     "luonnos_luotu": "create_draft",
}

PHASE_SECTION_NAME = "tietoryhmä"
PUBLIC_ATTRIBUTE = "tiedon julkisuus"  # kyllä/ei julkinen
HELP_TEXT = "ohje tiedon syöttäjälle"
HELP_LINK = "ohjeeseen liittyvä linkki"

# Project card related columns
CARD_SECTION_NAME = "tieto näkyy projektikortissa; projektikortin osio"
CARD_SECTION_LOCATION = "tieto näkyy projektikortissa; kenttien järjestys"
CARD_SECTION_DATE_FORMAT = "tieto näkyy projektikortissa; päivämäärän yhteydessä näkyvä teksti "
CARD_EXTERNAL_DOCUMENT_FIELDS = "tieto näkyy projektikortissa; valmiit dokumentit nimi ja linkki"
CARD_EXTERNAL_DOCUMENT_SECTION = "tieto näkyy projektikortissa; valmiit dokumentit ryhmien otsikot"
CARD_EXTERNAL_DOCUMENT_SECTION_INDEX = "tieto näkyy projektikortissa; valmiit dokumentit ryhmien järjestys"
CARD_SHOW_ON_MOBILE = "tieto näkyy näkymässä, joka on sovitettu mobiiliversioon"

# Overview-related columns
OVERVIEW_FILTER_NAME = "yleisnäkymä;\nsuodattimen nimi"
OVERVIEW_FILTERS = "yleisnäkymä;\nnäkymät joissa suodatinta käytetään"

# External API integration
EXT_DATA_SOURCE = "tiedon sijainti kaavoitus-api:ssa"
EXT_DATA_SOURCE_KEY = "tiedon polku kaavoitus-api:n paluudatassa"
EXT_DATA_KEY_ATTRIBUTE = "tiedon hakuavainkenttä kaavapinossa"
EXT_DATA_PARENT_KEY_ATTRIBUTE = "vanhemmalta peritty tiedon hakuavainkenttä kaavapinossa"
EXT_DATA_AD_SOURCE = "automaattisen yhteystiedon lähdekenttä"
EXT_DATA_AD_KEY = "automaattisen yhteystiedon hakuavain"

class OverviewViews(Enum):
    BY_SUBTYPE = "Kaavaprojektien jakauma"
    ON_MAP = "Kaavaprojektit kartalla"
    FLOOR_AREA = "Kaavoitettu kerrosala"


CALCULATIONS_COLUMN = "laskelmat"

ATTRIBUTE_FIELDSET = "projektitieto fieldset"

class Phases(Enum):
    START = "Käynnistys"
    PRINCIPLES = "Periaatteet"
    OAS = "OAS"
    DRAFT = "Luonnos"
    PROPOSAL = "Ehdotus"
    REVISED_PROPOSAL = "Tarkistettu ehdotus"
    APPROVAL = "Hyväksyminen"
    GOING_INTO_EFFECT = "Voimaantulo"


ATTRIBUTE_PHASE_COLUMNS = {
    Phases.START: ("käynnistys vaiheen otsikot", "käynnistys vaiheen alaotsikko", "käynnistys vaiheen järjestys"),
    Phases.PRINCIPLES: ("periaatteet vaiheen otsikot", "periaatteet vaiheen alaotsikko", "periaatteet vaiheen järjestys"),
    Phases.OAS: ("oas vaiheen otsikot", "oas vaiheen alaotsikko", "oas vaiheen järjestys"),
    Phases.DRAFT: ("luonnos vaiheen otsikot", "luonnos vaiheen alaotsikko", "luonnos vaiheen järjestys"),
    Phases.PROPOSAL: ("ehdotus vaiheen otsikot", "ehdotus vaiheen alaotsikko", "ehdotus vaiheen järjestys"),
    Phases.REVISED_PROPOSAL: ("tarkistettu ehdotus vaiheen otsikot", "tarkistettu ehdotus vaiheen alaotsikko", "tarkistettu ehdotus vaiheen järjestys"),
    Phases.APPROVAL: ("hyväksyminen vaiheen otsikot", "hyväksyminen vaiheen alaotsikko", "hyväksyminen vaiheen järjestys"),
    Phases.GOING_INTO_EFFECT: ("voimaantulo vaiheen otsikot", "voimaantulo vaiheen alaotsikko", "voimaantulo vaiheen järjestys"),
}

ATTRIBUTE_DEADLINE_SECTION_COLUMNS = {
    "owner": "vastuuhenkilön aikataulun muokkaus -näkymän osiot ja kenttien järjestys",
    "admin": "pääkäyttäjän päivämäärätietojen vahvistus -näkymän osiot ja kenttien järjestys",
}

ATTRIBUTE_CATEGORIZATION_COLUMNS = [
    ("tiedon luokitus käynnistys vaiheessa", Phases.START, []),
    ("tiedon luokitus käynnistys vaiheessa", Phases.START, [Phases.PRINCIPLES]),
    ("tiedon luokitus käynnistys vaiheessa", Phases.START, [Phases.PRINCIPLES, Phases.DRAFT]),
    ("tiedon luokitus käynnistys vaiheessa", Phases.START, [Phases.DRAFT]),

    ("tiedon luokitus periaatteet vaiheessa", Phases.PRINCIPLES, []),
    ("tiedon luokitus periaatteet vaiheessa", Phases.PRINCIPLES, [Phases.PRINCIPLES]),
    ("tiedon luokitus periaatteet vaiheessa", Phases.PRINCIPLES, [Phases.PRINCIPLES, Phases.DRAFT]),
    ("tiedon luokitus periaatteet vaiheessa", Phases.PRINCIPLES, [Phases.DRAFT]),

    ("tiedon luokitus oas vaiheessa, kun prosessi ei sisällä periaatteet-vaihetta", Phases.OAS, []),
    ("tiedon luokitus oas vaiheessa, kun prosessi sisältää periaate-vaiheen", Phases.OAS, [Phases.PRINCIPLES]),
    ("tiedon luokitus oas vaiheessa, kun prosessi sisältää periaate-vaiheen", Phases.OAS, [Phases.PRINCIPLES, Phases.DRAFT]),
    ("tiedon luokitus oas vaiheessa, kun prosessi ei sisällä periaatteet-vaihetta", Phases.OAS, [Phases.DRAFT]),

    ("tiedon luokitus luonnos vaiheessa, kun prosessi ei sisällä periaatteet-vaihetta", Phases.DRAFT, []),
    ("tiedon luokitus luonnos vaiheessa, kun prosessi sisältää periaate-vaiheen", Phases.DRAFT, [Phases.PRINCIPLES]),
    ("tiedon luokitus luonnos vaiheessa, kun prosessi sisältää periaate-vaiheen", Phases.DRAFT, [Phases.PRINCIPLES, Phases.DRAFT]),
    ("tiedon luokitus luonnos vaiheessa, kun prosessi ei sisällä periaatteet-vaihetta", Phases.DRAFT, [Phases.DRAFT]),

    ("tiedon luokitus ehdotus vaiheessa, kun prosessi ei sisällä periaatteet- eikä luonnos-vaihetta", Phases.PROPOSAL, []),
    ("tiedon luokitus ehdotus vaiheessa, kun prosessi sisältää periaatteet-vaiheen, mutta ei luonnos-vaihetta", Phases.PROPOSAL, [Phases.PRINCIPLES]),
    ("tiedon luokitus ehdotus vaiheessa, kun prosessi sisältää periaatteet- ja luonnos-vaiheen", Phases.PROPOSAL, [Phases.PRINCIPLES, Phases.DRAFT]),
    ("tiedon luokitus ehdotus vaiheessa, kun prosessi sisältää luonnos-vaiheen, mutta ei periaatteet-vaihetta", Phases.PROPOSAL, [Phases.DRAFT]),

    ("tiedon luokitus tarkistettu ehdotus vaiheessa, kun prosessi ei sisällä periaatteet- eikä luonnos-vaihetta", Phases.REVISED_PROPOSAL, []),
    ("tiedon luokitus tarkistettu ehdotus vaiheessa, kun prosessi sisältää periaatteet-vaiheen, mutta ei luonnos-vaihetta", Phases.REVISED_PROPOSAL, [Phases.PRINCIPLES]),
    ("tiedon luokitus tarkistettu ehdotus vaiheessa, kun prosessi sisältää periaatteet- ja luonnos-vaiheen", Phases.REVISED_PROPOSAL, [Phases.PRINCIPLES, Phases.DRAFT]),
    ("tiedon luokitus tarkistettu ehdotus vaiheessa, kun prosessi sisältää luonnos-vaiheen, mutta ei periaatteet-vaihetta", Phases.REVISED_PROPOSAL, [Phases.DRAFT]),

    ("tiedon luokitus hyväksyminen vaihe", Phases.APPROVAL, []),
    ("tiedon luokitus hyväksyminen vaihe", Phases.APPROVAL, [Phases.PRINCIPLES]),
    ("tiedon luokitus hyväksyminen vaihe", Phases.APPROVAL, [Phases.PRINCIPLES, Phases.DRAFT]),
    ("tiedon luokitus hyväksyminen vaihe", Phases.APPROVAL, [Phases.DRAFT]),

    ("tiedon luokitus voimaantulo vaihe", Phases.GOING_INTO_EFFECT, []),
    ("tiedon luokitus voimaantulo vaihe", Phases.GOING_INTO_EFFECT, [Phases.PRINCIPLES]),
    ("tiedon luokitus voimaantulo vaihe", Phases.GOING_INTO_EFFECT, [Phases.PRINCIPLES, Phases.DRAFT]),
    ("tiedon luokitus voimaantulo vaihe", Phases.GOING_INTO_EFFECT, [Phases.DRAFT]),
]

ATTRIBUTE_FLOOR_AREA_SECTION = "kerrosalatietojen muokkaus -näkymän osiot pääotsikot"
ATTRIBUTE_FLOOR_AREA_SECTION_MATRIX_ROW = "kerrosalatietojen muokkaus -näkymän alaotsikot"
ATTRIBUTE_FLOOR_AREA_SECTION_MATRIX_CELL = "kerrosalatietojen muokkaus -näkymän tietokenttien nimet"
EXPECTED_A1_VALUE = ATTRIBUTE_NAME

USER_PRIVILEGES = {
    "selaaja": "browse",
    "asiantuntija": "edit",
    "vastuuhenkilö": "create",
    "pääkäyttäjä": "admin",
}

HIGHLIGHT_GROUPS = {
    "Asiantuntijan kenttä": "Asiantuntijat",
    "Pääkäyttäjän kenttä": "Pääkäyttäjät",
}

KNOWN_SUBTYPES = ["XS", "S", "M", "L", "XL"]

DATA_RETENTION_PLANS = {
    "tieto tallennetaan pysyvästi": {
        "label": "tieto tallennetaan pysyvästi",
        "plan_type": DataRetentionPlan.TYPE_PERMANENT,
    },
    "prosessinaikainen": {
        "label": "prosessinaikainen",
        "plan_type": DataRetentionPlan.TYPE_PROCESSING,
    },
    "tieto poistuu 6 kk kuluttua, kun projekti on arkistoitu": {
        "label": "6 kk arkistoinnista",
        "plan_type": DataRetentionPlan.TYPE_CUSTOM,
        "custom_time": 6,
        "custom_time_unit": DataRetentionPlan.UNIT_MONTHS,
    },
}
DEFAULT_DATA_RETENTION_PLAN = DATA_RETENTION_PLANS["tieto tallennetaan pysyvästi"]

PROJECT_PHASES = {
    Phases.START.value: {
        "name": Phases.START.value,
        "color": "color-tram",
        "color_code": "#02d7a7",
        "list_prefix": "1",
    },  # None
    Phases.PRINCIPLES.value: {
        "name": Phases.PRINCIPLES.value,
        "color": 'color-green',
        "color_code": '#009246',
        "list_prefix": "XL",
    },
    Phases.OAS.value: {
        "name": Phases.OAS.value,
        "color": "color-summer",
        "color_code": "#ffc61e",
        "list_prefix": "2",
    },  # 01, 03
    Phases.DRAFT.value: {
        "name": Phases.DRAFT.value,
        "color": 'color-suomenlinna',
        "color_code": '#f5a3c7',
        "list_prefix": "XL",
    },
    Phases.PROPOSAL.value: {
        "name": Phases.PROPOSAL.value,
        "color": "color-metro",
        "color_code": "#fd4f00",
        "list_prefix": "3",
    },  # 02, 04
    Phases.REVISED_PROPOSAL.value: {
        "name": Phases.REVISED_PROPOSAL.value,
        "color": "color-bus",
        "color_code": "#0000bf",
        "list_prefix": "4",
    },  # 05, 07
    Phases.APPROVAL.value: {
        "name": Phases.APPROVAL.value,
        "color": "color-gold",
        "color_code": "#bd9650",
        "list_prefix": "5",
    },  # 06, 07 <- Kvsto
    Phases.GOING_INTO_EFFECT.value: {
        "name": Phases.GOING_INTO_EFFECT.value,
        "color": "color-fog",
        "color_code": "#9ec8eb",
        "list_prefix": "6",
    },
}

SUBTYPE_PHASES = {
    "XS": [
        Phases.START.value,
        Phases.OAS.value,
        Phases.PROPOSAL.value,
        Phases.REVISED_PROPOSAL.value,
        Phases.APPROVAL.value,
        Phases.GOING_INTO_EFFECT.value,
    ],
    "S": [
        Phases.START.value,
        Phases.OAS.value,
        Phases.PROPOSAL.value,
        Phases.REVISED_PROPOSAL.value,
        Phases.APPROVAL.value,
        Phases.GOING_INTO_EFFECT.value,
    ],
    "M": [
        Phases.START.value,
        Phases.OAS.value,
        Phases.PROPOSAL.value,
        Phases.REVISED_PROPOSAL.value,
        Phases.APPROVAL.value,
        Phases.GOING_INTO_EFFECT.value,
    ],
    "L": [
        Phases.START.value,
        Phases.OAS.value,
        Phases.PROPOSAL.value,
        Phases.REVISED_PROPOSAL.value,
        Phases.APPROVAL.value,
        Phases.GOING_INTO_EFFECT.value,
    ],
    "XL": [
        Phases.START.value,
        Phases.PRINCIPLES.value,
        Phases.OAS.value,
        Phases.DRAFT.value,
        Phases.PROPOSAL.value,
        Phases.REVISED_PROPOSAL.value,
        Phases.APPROVAL.value,
        Phases.GOING_INTO_EFFECT.value,
    ],
}

SUBTYPE_PHASE_METADATA = {
    "XS": {
        Phases.START.value: {"default_end_weeks_delta": 8},
        Phases.OAS.value: {"default_end_weeks_delta": 12},
        Phases.PROPOSAL.value: {"default_end_weeks_delta": 24},
        Phases.REVISED_PROPOSAL.value: {"default_end_weeks_delta": 10},
        Phases.APPROVAL.value: {"default_end_weeks_delta": 12},
        Phases.GOING_INTO_EFFECT.value: {"default_end_weeks_delta": 6},
    },
    "S": {
        Phases.START.value: {"default_end_weeks_delta": 8},
        Phases.OAS.value: {"default_end_weeks_delta": 12},
        Phases.PROPOSAL.value: {"default_end_weeks_delta": 24},
        Phases.REVISED_PROPOSAL.value: {"default_end_weeks_delta": 10},
        Phases.APPROVAL.value: {"default_end_weeks_delta": 12},
        Phases.GOING_INTO_EFFECT.value: {"default_end_weeks_delta": 6},
    },
    "M": {
        Phases.START.value: {"default_end_weeks_delta": 8},
        Phases.OAS.value: {"default_end_weeks_delta": 12},
        Phases.PROPOSAL.value: {"default_end_weeks_delta": 24},
        Phases.REVISED_PROPOSAL.value: {"default_end_weeks_delta": 10},
        Phases.APPROVAL.value: {"default_end_weeks_delta": 12},
        Phases.GOING_INTO_EFFECT.value: {"default_end_weeks_delta": 6},
    },
    "L": {
        Phases.START.value: {"default_end_weeks_delta": 12},
        Phases.OAS.value: {"default_end_weeks_delta": 16},
        Phases.PROPOSAL.value: {"default_end_weeks_delta": 36},
        Phases.REVISED_PROPOSAL.value: {"default_end_weeks_delta": 14},
        Phases.APPROVAL.value: {"default_end_weeks_delta": 12},
        Phases.GOING_INTO_EFFECT.value: {"default_end_weeks_delta": 6},
    },
    "XL": {
        Phases.START.value: {"default_end_weeks_delta": 12},
        Phases.PRINCIPLES.value: {"default_end_weeks_delta": 0},
        Phases.OAS.value: {"default_end_weeks_delta": 16},
        Phases.DRAFT.value: {"default_end_weeks_delta": 0},
        Phases.PROPOSAL.value: {"default_end_weeks_delta": 36},
        Phases.REVISED_PROPOSAL.value: {"default_end_weeks_delta": 14},
        Phases.APPROVAL.value: {"default_end_weeks_delta": 12},
        Phases.GOING_INTO_EFFECT.value: {"default_end_weeks_delta": 6},
    },
}

VALUE_TYPES = {
    "AD-tunnukset": Attribute.TYPE_USER,
    "AD-tunnukset, koko organisaatio": Attribute.TYPE_PERSONNEL,
    "automaatinen (teksti), kun valitaan henkilö": Attribute.TYPE_SHORT_STRING,
    "automaattinen (desimaaliluku), tieto tulee kaavan tietomallista": Attribute.TYPE_DECIMAL,
    "automaattinen (kokonaisluku), jonka Kaavapino laskee": Attribute.TYPE_INTEGER,
    "automaattinen (kokonaisluku), kun projekti luodaan": Attribute.TYPE_INTEGER,
    "automaattinen (kokonaisluku), kun projekti luodaan, kokonaisluku": Attribute.TYPE_INTEGER,
    "automaattinen (kokonaisluku), tieto tulee Factasta": Attribute.TYPE_INTEGER,
    "automaattinen (kokonaisluku), tieto tulee kaavan tietomallista": Attribute.TYPE_INTEGER,
    "automaattinen (pvm)": Attribute.TYPE_DATE,
    "automaattinen (pvm), mutta päivämäärää voi muuttaa": Attribute.TYPE_DATE,
    "automaattinen (spatiaalinen), tieto tulee kaavan tietomallista": Attribute.TYPE_GEOMETRY,
    "automaattinen (teksti), jonka Kaavapino muodostaa": Attribute.TYPE_SHORT_STRING,
    "automaattinen (teksti), kun projekti luodaan": Attribute.TYPE_SHORT_STRING,
    "automaattinen (teksti), kun valitaan vastuuyksikkö": Attribute.TYPE_SHORT_STRING,
    "automaattinen (teksti), tieto tulee Factasta": Attribute.TYPE_SHORT_STRING,
    "automaattinen (valinta)": Attribute.TYPE_BOOLEAN,
    "automaattinen (valinta), kun projekti luodaan": Attribute.TYPE_LONG_STRING,
    "Desimaaliluvun syöttö": Attribute.TYPE_DECIMAL,
    "fieldset": Attribute.TYPE_FIELDSET,
    "info_fieldset": Attribute.TYPE_INFO_FIELDSET,
    "Kokonaisluvun syöttö.": Attribute.TYPE_INTEGER,
    "Kuvan lataaminen.": Attribute.TYPE_IMAGE,
    "Kyllä/Ei": Attribute.TYPE_BOOLEAN,
    "Kyllä/Ei/Tieto puuttuu": Attribute.TYPE_BOOLEAN,
    "Linkin liittäminen.": Attribute.TYPE_LINK,
    "Lyhyen tekstin syöttö.": Attribute.TYPE_RICH_TEXT_SHORT,
    "Numerosarjan syöttö.": Attribute.TYPE_SHORT_STRING,
    "Pitkän tekstin syöttö.": Attribute.TYPE_RICH_TEXT,
    "Päivämäärän valinta.": Attribute.TYPE_DATE,
    "Valinta (1) pudotusvalikosta.": Attribute.TYPE_CHOICE,
    "Valinta (1) valintapainikkeesta.": Attribute.TYPE_CHOICE,
    "Valinta (1-2) painikkeesta.": Attribute.TYPE_CHOICE,
    "Valinta (1-x) pudotusvalikosta.": Attribute.TYPE_CHOICE,
    "Valintaruutu.": Attribute.TYPE_BOOLEAN,
    "Vuoden valinta pudotusvalikosta": Attribute.TYPE_INTEGER,
    "Valinta pudotusvalikosta. (readonly)": Attribute.TYPE_CHOICE,
    "Valintaruutu. (readonly)": Attribute.TYPE_BOOLEAN,
    "Päivämäärä (readonly)": Attribute.TYPE_DATE,
    "Lyhyt teksti (readonly)": Attribute.TYPE_RICH_TEXT_SHORT,
    "Kyllä/Ei (readonly)": Attribute.TYPE_BOOLEAN,
    "Kyllä/Ei/Tieto puuttuu (readonly)": Attribute.TYPE_BOOLEAN,
    "Kokonaisluku (readonly)": Attribute.TYPE_INTEGER,
    "Vuosiluku ilman tuhaterotinta": Attribute.TYPE_INTEGER,
    "Kokonaisluku ilman tuhaterotinta": Attribute.TYPE_INTEGER,
    "muotoilematon tekstikenttä": Attribute.TYPE_SHORT_STRING,
}

DISPLAY_TYPES = {
    "Valinta (1) pudotusvalikosta.": Attribute.DISPLAY_DROPDOWN,
    "Valinta (1-x) pudotusvalikosta.": Attribute.DISPLAY_DROPDOWN,
    "Vuoden valinta pudotusvalikosta": Attribute.DISPLAY_DROPDOWN,
    "Valintaruutu.": Attribute.DISPLAY_CHECKBOX,
    "Valinta pudotusvalikosta. (readonly)": Attribute.DISPLAY_READONLY,
    "Valintaruutu. (readonly)": Attribute.DISPLAY_READONLY_CHECKBOX,
    "Päivämäärä (readonly)": Attribute.DISPLAY_READONLY,
    "Lyhyt teksti (readonly)": Attribute.DISPLAY_READONLY,
    "Kyllä/Ei (readonly)": Attribute.DISPLAY_READONLY_CHECKBOX,
    "Kyllä/Ei/Tieto puuttuu (readonly)": Attribute.DISPLAY_READONLY_CHECKBOX,
    "Kokonaisluku (readonly)": Attribute.DISPLAY_READONLY,
    "automaatinen (teksti), kun valitaan henkilö": Attribute.DISPLAY_READONLY,
    "automaattinen (desimaaliluku), tieto tulee kaavan tietomallista": Attribute.DISPLAY_READONLY,
    "automaattinen (kokonaisluku), jonka Kaavapino laskee": Attribute.DISPLAY_READONLY,
    "automaattinen (kokonaisluku), kun projekti luodaan": Attribute.DISPLAY_READONLY,
    "automaattinen (kokonaisluku), kun projekti luodaan, kokonaisluku": Attribute.DISPLAY_READONLY,
    "automaattinen (kokonaisluku), tieto tulee Factasta": Attribute.DISPLAY_READONLY,
    "automaattinen (kokonaisluku), tieto tulee kaavan tietomallista": Attribute.DISPLAY_READONLY,
    "automaattinen (pvm)": Attribute.DISPLAY_READONLY,
    "automaattinen (spatiaalinen), tieto tulee kaavan tietomallista": Attribute.DISPLAY_READONLY,
    "automaattinen (teksti), jonka Kaavapino muodostaa": Attribute.DISPLAY_READONLY,
    "automaattinen (teksti), kun projekti luodaan": Attribute.DISPLAY_READONLY,
    "automaattinen (teksti), kun valitaan vastuuyksikkö": Attribute.DISPLAY_READONLY,
    "automaattinen (teksti), tieto tulee Factasta": Attribute.DISPLAY_READONLY,
    "automaattinen (valinta)": Attribute.DISPLAY_READONLY_CHECKBOX,
    "automaattinen (valinta), kun projekti luodaan": Attribute.DISPLAY_READONLY,
    "Vuosiluku ilman tuhaterotinta": Attribute.DISPLAY_SIMPLE_INTEGER,
    "Kokonaisluku ilman tuhaterotinta": Attribute.DISPLAY_SIMPLE_INTEGER,
}


class AttributeImporterException(Exception):
    pass


class AttributeUpdater:
    def __init__(self, options=None):
        self.options = options

    @transaction.atomic
    def run(self):
        logger.info(f"Admin: Starting to update attribute identifiers")
        file = self.options['filename'].read().decode('utf-8')
        reader = csv.reader(StringIO(file))

        for i, row in enumerate(reader):
            if i == 0:
                continue

            old_identifier = row[0]
            new_identifier = row[1]

            if not old_identifier or not new_identifier:
                logger.info(f"Invalid input data: Identifier to replace: {old_identifier} New identifier: {new_identifier}")
                continue

            logger.info(f"Updating attributes: {old_identifier}->{new_identifier}")
            attributes = Attribute.objects.filter(
                Q(identifier=old_identifier)
                | Q(key_attribute_path__contains=old_identifier)
                | Q(related_fields__contains=[old_identifier])
                | Q(visibility_conditions__0__contains={"variable": old_identifier})
            )
            for attr in attributes:
                logger.info(f"    Updating attribute {attr.id} {attr.identifier}")
                if attr.identifier == old_identifier:
                    attr.identifier = new_identifier
                    logger.info(f"        Updated Attribute.identifier")

                if attr.key_attribute_path and old_identifier in attr.key_attribute_path:
                    attr.key_attribute_path = attr.key_attribute_path.replace(
                        old_identifier,
                        new_identifier
                    )
                    logger.info(f"        Updated Attribute.key_attribute_path")

                if attr.related_fields and old_identifier in attr.related_fields:
                    updated = []
                    for i, rel_f in enumerate(attr.related_fields):
                        if rel_f == old_identifier:
                            attr.related_fields[i] = new_identifier
                            logger.info(f"        Updated Attribute.related_fields")

                for i, vis_con in enumerate(attr.visibility_conditions):
                    if vis_con.get("variable") == old_identifier:
                        attr.visibility_conditions[i]["variable"] = new_identifier
                        logger.info(f"        Updated Attribute.visibility_conditions")


                attr.save()


            logger.info(f"Updating project attribute_data: {old_identifier}->{new_identifier}")
            for proj in Project.objects.all():
                attr_data = json.dumps(proj.attribute_data)
                if old_identifier in attr_data:
                    logger.info(f"    Updating project {proj.id}: {old_identifier}->{new_identifier}")
                    attr_data = attr_data.replace(old_identifier, new_identifier)
                    proj.attribute_data = json.loads(attr_data)
                    proj.save()



class AttributeUpdaterException(Exception):
    pass


class AttributeImporter:
    """Import attributes and project phase sections for asemakaava project type from the given Excel."""

    def __init__(self, options=None):
        self.options = options
        self.workbook = None

    def _open_workbook(self, filename):
        try:
            return load_workbook(filename, read_only=True, data_only=True)
        except FileNotFoundError as e:
            raise AttributeImporterException(e)

    def _extract_data_from_workbook(self, workbook):
        try:
            sheet = workbook[self.options.get("sheet") or DEFAULT_SHEET_NAME]
        except KeyError as e:
            raise AttributeImporterException(e)

        primary_sheet_value = sheet["A1"].value
        if primary_sheet_value and primary_sheet_value.lower() != EXPECTED_A1_VALUE:
            raise AttributeImporterException(
                "This does not seem to be a valid attribute sheet."
            )

        return self._rows_for_sheet(sheet, 1)

    def _rows_for_sheet(self, sheet, start=0):
        data = []

        for row in list(sheet.iter_rows())[start:]:
            if not row[0].value:
                break
            data.append([col.value for col in row])

        return data

    def _set_row_indexes(self, header_row: Iterable[str]):
        """Determine index number for all columns."""
        self.column_index: dict = {}

        for index, column in enumerate(header_row):
            if column:
                self.column_index[column.lower()] = index

    def _check_if_row_valid(self, row: Sequence) -> bool:
        """Check if the row has all required data.

        For importing a field the requirements are:
        - Name
        - Type
        """
        name = row[self.column_index[ATTRIBUTE_NAME]]
        attr_type = row[self.column_index[ATTRIBUTE_TYPE]]

        if name and attr_type:
            return True

        logger.info(f"Field {name} is not valid, it will not be imported.")
        return False

    def _row_part_of_fieldset(self, row: Sequence) -> bool:
        """Check if the row should be a part of a fieldset."""
        if ATTRIBUTE_FIELDSET not in self.column_index:
            return False
        fieldset_value = row[self.column_index[ATTRIBUTE_FIELDSET]]
        return bool(fieldset_value)

    def _get_identifier_for_value(self, value: str) -> str:
        identifier = create_identifier(value.strip(" \t:."))
        return truncate_identifier(identifier, length=IDENTIFIER_MAX_LENGTH)

    def _get_attribute_row_identifier(self, row: Sequence) -> str:
        predefined_identifier = row[self.column_index[ATTRIBUTE_IDENTIFIER]]
        if predefined_identifier:
            predefined_identifier = predefined_identifier.strip()
            if not check_identifier(predefined_identifier):
                raise ValueError(
                    f"The identifier '{predefined_identifier}' is not a proper slug value"
                )
            return predefined_identifier

        logger.info(
            f"\nCreating identifier for '{row[self.column_index[ATTRIBUTE_NAME]]}'"
        )
        return self._get_identifier_for_value(row[self.column_index[ATTRIBUTE_NAME]])

    def _create_attributes(self, rows: Iterable[Sequence[str]]):
        def parse_condition(condition):
            condition = re.split(r"\s+(not in|in|\=\=|\!\=|\>|\<)+\s+", condition)

            if len(condition) == 1:
                negate = condition[0][0] == "!"
                condition = [
                    condition[0][1:] if negate else condition[0],
                    "!=" if negate else "==",
                    True,
                ]
                value = condition[2]
                value_type = "boolean"

            else:
                value = condition[2]
                if value[0] == '"':
                    value = value[1:]

                if value[-1] == '"':
                    value = value[:-1]

                if value[0] == "[" and value[-1] == "]":
                    try:
                        [int(i) for i in re.split(r",\s+", value[1:-1])]
                        value_type = "list<number>"
                    except ValueError:
                        value_type = "list<string>"
                else:
                    try:
                        int(value)
                        value_type = "number"
                    except ValueError:
                        value_type = "string"

            return {
                "variable": condition[0],
                "operator": condition[1],
                "comparison_value": value,
                "comparison_value_type": value_type,
            }

        def parse_autofill_rule(rule, type):
            if rule == "ei":
                return None

            # TODO: make a more general implementation for including fields if cases become more complex
            variables = re.findall(r"^\{\{(.*)\}\}", rule)
            thens = re.findall(r"\{%+\sif.*?%\}\s*(.*?)\s*\{% endif %\}", rule)
            conditions = re.findall(r"\{%\s*if\s*(.*?)\s*%\}.*?\{%\s*endif\s*%\}", rule)

            branches = []

            for (then, condition) in zip(thens, conditions):
                if type == Attribute.TYPE_CHOICE:
                    then = self._get_identifier_for_value(str(then))
                elif type == Attribute.TYPE_BOOLEAN and then == "kyllä":
                    then = True
                elif type == Attribute.TYPE_BOOLEAN and then == "ei":
                    then = False

                new_branches = [
                    {
                        "variables": variables,
                        "conditions": [
                            parse_condition(condition_and)
                            for condition_and
                            in re.split(r"\sand+\s", condition_or)
                        ],
                        "then_branch": then,
                        "else_branch": None
                    }
                    for condition_or in re.split(r"\sor+\s", condition)
                ]
                branches += new_branches

            return branches

        def parse_autofill_readonly(rule):
            if rule == "ei" or \
                (rule and rule.startswith("Automaattiseti muodostunutta tietoa ei voi muokata")):
                return True

            return False

        logger.info("\nCreating attributes...")

        existing_attribute_ids = set(
            Attribute.objects.all().values_list("identifier", flat=True)
        )

        imported_attribute_ids = set()
        created_attribute_count = 0
        updated_attribute_count = 0
        created_choices_count = 0
        for row in rows:
            identifier = self._get_attribute_row_identifier(row)

            name = row[self.column_index[ATTRIBUTE_NAME]].strip(" \t:.")

            value_type_string = row[self.column_index[ATTRIBUTE_TYPE]]
            value_type = (
                VALUE_TYPES.get(value_type_string.strip())
                if value_type_string
                else None
            )
            display = (
                DISPLAY_TYPES.get(value_type_string.strip(), None)
                if value_type_string
                else None
            )
            visibility_row = row[self.column_index[ATTRIBUTE_RULE_CONDITIONAL_VISIBILITY]] or ""
            ifs = re.findall(
                r"\{%\s*if\s*(.*?)\s*%\}",
                visibility_row,
            )
            conditions = []

            for if_condition in ifs:
                conditions += re.split(r"\s+or\s+", if_condition)

            if len(re.findall(r"ei\s*\{%\s*endif\s*%\}", visibility_row)):
                visibility_conditions = []
                hide_conditions = [
                    parse_condition(condition)
                    for condition in conditions
                ]
            else:
                visibility_conditions = [
                    parse_condition(condition)
                    for condition in conditions
                ]
                hide_conditions = []

            unit = row[self.column_index[ATTRIBUTE_UNIT]] or None

            broadcast_changes = (
                row[self.column_index[ATTRIBUTE_BROADCAST_CHANGES]] == "kyllä"
            )

            try:
                data_retention_plan = DATA_RETENTION_PLANS[
                    row[self.column_index[ATTRIBUTE_DATA_RETENTION]]
                ]
            except KeyError:
                data_retention_plan = DEFAULT_DATA_RETENTION_PLAN

            data_retention_plan, _ = DataRetentionPlan.objects.update_or_create(
                label=data_retention_plan["label"],
                defaults=data_retention_plan,
            )

            multiple_choice = row[self.column_index[ATTRIBUTE_MULTIPLE_CHOICE]] in [
                "kyllä",
                "fieldset voi toistua",
            ]
            try:
                character_limit = int(row[self.column_index[ATTRIBUTE_CHARACTER_LIMIT]])
            except (TypeError, ValueError):
                character_limit = None

            try:
                validation_regex = row[self.column_index[ATTRIBUTE_VALIDATION_REGEX]]
                if validation_regex:
                    re.compile(validation_regex)  # Check that given regex is valid
            except re.error:
                validation_regex = None

            try:
                help_text = row[self.column_index[HELP_TEXT]].strip()
            except (IndexError, AttributeError):
                help_text = ""

            try:
                help_link = row[self.column_index[HELP_LINK]].strip()
            except (IndexError, AttributeError):
                help_link = None

            is_public = row[self.column_index[PUBLIC_ATTRIBUTE]] == "kyllä"
            is_required = row[self.column_index[ATTRIBUTE_REQUIRED]] == "kyllä"
            is_searchable = row[self.column_index[ATTRIBUTE_SEARCHABLE]] == "kyllä"
            # TODO: ask for a dedicated column for uniqueness at some point
            is_unique = row[self.column_index[ATTRIBUTE_ERROR]] in ATTRIBUTE_ERROR_UNIQUE
            placeholder_text = row[self.column_index[ATTRIBUTE_PLACEHOLDER]]
            assistive_text = row[self.column_index[ATTRIBUTE_ASSISTIVE_TEXT]]
            error_text = row[self.column_index[ATTRIBUTE_ERROR_TEXT]]
            error_message = row[self.column_index[ATTRIBUTE_ERROR]]
            static_property = STATIC_ATTRIBUTES_MAPPING.get(
                row[self.column_index[ATTRIBUTE_IDENTIFIER]]
            )

            try:
                highlight_group = Group.objects.get(name=HIGHLIGHT_GROUPS[
                    row[self.column_index[ATTRIBUTE_HIGHLIGHT_GROUP]]
                ])
            except Group.DoesNotExist:
                raise Group.DoesNotExist(
                    "Default group(s) not found, try running create_default_groups_and_mappings management command first"
                )
            except KeyError:
                highlight_group = None

            related_fields = re.findall(
                r"\{\{(.*?)\}\}",
                row[self.column_index[ATTRIBUTE_RELATED_FIELDS]] or ""
            )
            linked_fields_string = row[self.column_index[ATTRIBUTE_LINKED_FIELDS]] or ""
            linked_fields = linked_fields_string.split(";") if ";" in linked_fields_string \
                else [linked_fields_string]

            # autofill
            try:
                autofill_rule = parse_autofill_rule(
                    row[self.column_index[ATTRIBUTE_RULE_AUTOFILL]],
                    value_type,
                )
                autofill_readonly = parse_autofill_readonly(
                    row[self.column_index[ATTRIBUTE_RULE_AUTOFILL_READONLY]]
                )
            except TypeError:
                autofill_rule = None
                autofill_readonly = None


            updates_autofill = row[self.column_index[ATTRIBUTE_RULE_UPDATE_AUTOFILL]] == "kyllä"


            if not value_type:
                logger.warning(
                    f'Unidentified value type "{value_type_string}", defaulting to short string'
                )
                value_type = Attribute.TYPE_SHORT_STRING

            generated, calculations = self._get_generated_calculations(row)

            owner_editable = len(re.findall(
                "Projektin vastuuhenkilö",
                row[self.column_index[ATTRIBUTE_EDIT_PRIVILEGE]] or "",
            )) > 0
            edit_privilege = row[self.column_index[ATTRIBUTE_EDIT_PRIVILEGE]]

            if edit_privilege == "automaattinen tieto, jota ei voi muokata":
                edit_privilege = None
            elif edit_privilege:
                privilege_list = re.split(r",\s*", edit_privilege.lower())
                for match_text, privilege in USER_PRIVILEGES.items():
                    if match_text in privilege_list:
                        edit_privilege = privilege
                        break

            field_roles = row[self.column_index[ATTRIBUTE_FIELD_ROLE]]
            if field_roles == "automaattinen tieto, jota ei voi muokata":
                field_roles = None
            field_subroles = row[self.column_index[ATTRIBUTE_FIELD_SUBROLE]]
            if field_subroles == "automaattinen tieto, jota ei voi muokata":
                field_subroles = None

            data_source = row[self.column_index[EXT_DATA_SOURCE]]
            data_source_key = row[self.column_index[EXT_DATA_SOURCE_KEY]]
            key_attribute_path = row[self.column_index[EXT_DATA_PARENT_KEY_ATTRIBUTE]]
            ad_data_key = row[self.column_index[EXT_DATA_AD_KEY]]

            attribute, created = Attribute.objects.update_or_create(
                identifier=identifier,
                defaults={
                    "name": name,
                    "value_type": value_type,
                    "display": display,
                    "visibility_conditions": visibility_conditions,
                    "hide_conditions": hide_conditions,
                    "help_text": help_text,
                    "help_link": help_link,
                    "public": is_public,
                    "required": is_required,
                    "searchable": is_searchable,
                    "multiple_choice": multiple_choice,
                    "data_retention_plan": data_retention_plan,
                    "character_limit": character_limit,
                    "validation_regex": validation_regex,
                    "placeholder_text": placeholder_text,
                    "assistive_text": assistive_text,
                    "error_text": error_text,
                    "unique": is_unique,
                    "error_message": error_message,
                    "generated": generated,
                    "calculations": calculations,
                    "related_fields": related_fields,
                    "linked_fields": linked_fields,
                    "unit": unit,
                    "broadcast_changes": broadcast_changes,
                    "autofill_rule": autofill_rule,
                    "autofill_readonly": autofill_readonly,
                    "updates_autofill": updates_autofill,
                    "highlight_group": highlight_group,
                    "static_property": static_property,
                    "owner_editable": owner_editable,
                    "edit_privilege": edit_privilege,
                    "owner_viewable": True,
                    "view_privilege": "browse",
                    "data_source": data_source,
                    "data_source_key": data_source_key,
                    "key_attribute_path": key_attribute_path,
                    "ad_data_key": ad_data_key,
                    "field_roles": field_roles,
                    "field_subroles": field_subroles,
                },
            )
            if created:
                created_attribute_count += 1
            else:
                updated_attribute_count += 1
            imported_attribute_ids.add(identifier)

            choices_ref = row[self.column_index[ATTRIBUTE_CHOICES_REF]]
            if choices_ref:
                created_choices_count += self._create_attribute_choices(attribute, row)
            else:
                AttributeValueChoice.objects.filter(attribute=attribute).delete()


            if created:
                logger.info(f"Created {attribute}")

        # Remove any attributes that was not imported
        old_attribute_ids = existing_attribute_ids - imported_attribute_ids
        logger.info(f"Old Attributes: {old_attribute_ids}")
        for old_id in sorted(old_attribute_ids, reverse=True):
            attr = Attribute.objects.get(identifier=old_id)
            deadlines = attr.deadline.all()

            # Delete related comments
            attr.comments.all().delete()

            # Remove relations to deadlines
            for dl in deadlines:
                dl.attribute = None
                dl.save()

            attr.delete()

        return {
            "created": created_attribute_count,
            "updated": updated_attribute_count,
            "deleted": len(old_attribute_ids),
            "choices": created_choices_count,
        }

    def _create_attribute_key_relations(self, rows):
        # remove all old relations
        for attr in Attribute.objects.filter(key_attribute__isnull=False):
            attr.key_attribute = None
            attr.save()

        for attr in Attribute.objects.filter(ad_key_attribute__isnull=False):
            attr.ad_key_attribute = None
            attr.save()

        AttributeAutoValue.objects.all().delete()

        for row in rows:
            identifier = self._get_attribute_row_identifier(row)
            key_identifier = row[self.column_index[EXT_DATA_KEY_ATTRIBUTE]]
            ad_key_identifier = row[self.column_index[EXT_DATA_AD_SOURCE]]
            auto_value_key_identifier = \
                row[self.column_index[ATTRIBUTE_AUTO_VALUE_KEY_FIELD]]

            auto_value_mapping = {}
            if row[self.column_index[ATTRIBUTE_AUTO_VALUE_MAPPING]]:
                for item in re.split(
                    r',\s*',
                    row[self.column_index[ATTRIBUTE_AUTO_VALUE_MAPPING]],
                ):
                    try:
                        [key, value] = \
                            [i.strip('"') for i in re.split(r':\s*', item)]
                        auto_value_mapping[key] = value
                    except ValueError:
                        continue

            attr = Attribute.objects.get(identifier=identifier)

            try:
                key_attr = Attribute.objects.get(identifier=key_identifier)
                attr.key_attribute = key_attr
            except Attribute.DoesNotExist:
                pass

            try:
                ad_key_attr = Attribute.objects.get(identifier=ad_key_identifier)
                attr.ad_key_attribute = ad_key_attr
            except Attribute.DoesNotExist:
                pass

            try:
                auto_value_key_attr = Attribute.objects.get(
                    identifier=auto_value_key_identifier,
                )
                auto_attr = AttributeAutoValue.objects.create(
                    value_attribute = attr,
                    key_attribute = auto_value_key_attr,
                )
                for (key, value) in auto_value_mapping.items():
                    if auto_attr.key_attribute.value_choices.count():
                        try:
                            key = auto_attr.key_attribute.value_choices \
                                .get(value=key).identifier
                        except AttributeValueChoice.DoesNotExist:
                            pass

                    AttributeAutoValueMapping.objects.create(
                        auto_attr=auto_attr,
                        key_str=key,
                        value_str=value,
                    )
            except Attribute.DoesNotExist:
                pass

            attr.save()

    def _get_generated_calculations(self, row):
        calculations_string = row[self.column_index[CALCULATIONS_COLUMN]]
        if calculations_string in [None, 'ei']:
            return False, None

        # Splits the string when a word or +, -, *, / operators is found
        calculations = re.findall(r"([\w]+|[\+\-\*/])+", calculations_string)

        return True, calculations

    def _create_attribute_choices(self, attribute, row) -> int:
        AttributeValueChoice.objects.filter(attribute=attribute).delete()

        created_choices_count = 0
        choices_rows = self._rows_for_sheet(self.workbook[CHOICES_SHEET_NAME])
        choices_ref = row[self.column_index[ATTRIBUTE_CHOICES_REF]]

        try:
            column_index = choices_rows[0].index(choices_ref)
        except ValueError:
            choices_rows = choices_ref.split("/")
            column_index = -1

        for index, choice_row in enumerate(choices_rows):
            if column_index < 0:
                choice = choice_row
            elif index == 0:
                continue
            else:
                choice = choice_row[column_index]

            identifier = self._get_identifier_for_value(str(choice))
            if not choice:
                break

            try:
                _, created = AttributeValueChoice.objects.update_or_create(
                    attribute=attribute,
                    identifier=identifier,
                    defaults={
                        "index": index,
                        "value": choice,
                    }
                )
            except IntegrityError:
                logger.warning(f'Duplicate choice "{choice} ({identifier})" for {attribute}, ignoring row')

            if created:
                created_choices_count += 1


        return created_choices_count

    def _create_fieldset_links(self, subtype, rows: Iterable[Sequence[str]]):
        logger.info("\nCreating fieldsets...")

        if ATTRIBUTE_FIELDSET not in self.column_index:
            logger.warning(f'Fieldset column "{ATTRIBUTE_FIELDSET}" missing: Skipping')
            return

        fieldset_map = defaultdict(list)
        phases = ProjectPhase.objects.filter(project_subtype=subtype)

        # Map out the link that need to be created
        for row in rows:
            fieldset_attr = row[self.column_index[ATTRIBUTE_FIELDSET]]
            if not fieldset_attr:
                continue

            phase_indices = []
            for phase in phases:
                location = self._get_attribute_locations(row, phase.name)
                if location is None:
                    continue

                try:
                    index = location["child_locations"][-1]
                except IndexError:
                    index = None

                if index is not None:
                    phase_indices.append((phase, index))

            attr_id = self._get_attribute_row_identifier(row)
            fieldset_map[fieldset_attr].append((attr_id, phase_indices))

        # Create the links
        for source_id in fieldset_map:
            source = Attribute.objects.get(identifier=source_id)

            for target_id, phase_indices in fieldset_map[source_id]:
                target = Attribute.objects.get(identifier=target_id)
                fsa, _ = FieldSetAttribute.objects.get_or_create(
                    attribute_source=source, attribute_target=target
                )
                for phase, index in phase_indices:
                    ProjectPhaseFieldSetAttributeIndex.objects.create(
                        index=index,
                        phase=phase,
                        attribute=fsa,
                    )
                logger.info(f"Created {fsa}")

    def _validate_generated_attributes(self):
        """
        Additional validation for attributes which generates their values

        These check can no be done in the model since the attributes in the
        calculations might not exist when the attribute is created, however
        in the importer we want to make sure that the values entered in the
        XLSX file does actually exist as that should be the case at all times.
        """

        generated_attributes = Attribute.objects.filter(generated=True)

        for attribute in generated_attributes:
            calculations = attribute.calculations
            error = False

            calculation_attributes = Attribute.objects.filter(
                identifier__in=attribute.calculation_attribute_identifiers
            )
            if calculation_attributes.count() != len(
                attribute.calculation_attribute_identifiers
            ):
                logger.warning(
                    f"Could not find the attributes given in calculation. "
                    f"({calculation_attributes.count()}/{len(attribute.calculation_attribute_identifiers)})."
                    f"Error in {attribute.identifier} with calculation {calculations}."
                )
                error = True

            if not all(
                attribute.value_type in VALID_ATTRIBUTE_CALCULATION_TYPES
                for attribute in calculation_attributes
            ):
                logger.warning(
                    f"Calculation attributes are not valid number attributes. "
                    f"Error in {attribute.identifier} with calculation {calculations}."
                )
                error = True

            if error:
                raise Exception(f"Could not add attribute {attribute.identifier}")

    def _get_attribute_locations(self, row, phase_name):
        for phase, (column_label, column_ingress, column_location) in ATTRIBUTE_PHASE_COLUMNS.items():
            if phase.value == phase_name:  # e.g. Käynnistys
                label = row[self.column_index[column_label]]
                ingress = row[self.column_index[column_ingress]]
                location = row[self.column_index[column_location]]
                try:
                    locations = [
                        # location format is section.field(set).field
                        # where : acts as optional decimal separator
                        int(float(".".join(loc.split(':'))) * 10000)
                        for loc in location.split(".")
                    ]
                except (ValueError, AttributeError):
                    return None

                return {
                    "label": label,
                    "ingress": ingress,
                    "section_location": locations[0] if locations else None,
                    "field_location": locations[1] if len(locations) > 1 else 0,
                    "child_locations": locations[2:],
                }

        return None

    def _create_document_link_sections(self, rows):
        logger.info("\nReplacing external document link sections...")
        DocumentLinkFieldSet.objects.all().delete()

        for row in rows:
            section_name = row[self.column_index[CARD_EXTERNAL_DOCUMENT_SECTION]]
            section_index = row[self.column_index[CARD_EXTERNAL_DOCUMENT_SECTION_INDEX]]

            if not (section_name and section_index):
                continue

            identifiers = row[self.column_index[CARD_EXTERNAL_DOCUMENT_FIELDS]].split(";")

            if len(identifiers) != 3:
                logger.warning(f"Invalid attribute identifier formatting {fieldset.identifier}, ignoring")
                continue

            section, __ = DocumentLinkSection.objects.update_or_create(
                name=section_name,
                defaults={
                    "index": int(section_index),
                },
            )
            fieldset = Attribute.objects.get(
                identifier=row[self.column_index[ATTRIBUTE_IDENTIFIER]]
            )
            try:
                [name, custom_name, link] = [
                    Attribute.objects.get(identifier=identifier)
                    if identifier else None
                    for identifier in identifiers
                ]
            except Attribute.DoesNotExist:
                logger.warning(f"Invalid document link attributes specified for {fieldset.identifier}, ignoring")

            DocumentLinkFieldSet.objects.create(
                section=section,
                fieldset_attribute=fieldset,
                document_name_attribute=name,
                document_custom_name_attribute=custom_name,
                document_link_attribute=link,
            )


    def _create_card_sections(self, rows):
        logger.info("\nReplacing project card sections...")
        ProjectCardSection.objects.all().delete()
        ProjectCardSectionAttribute.objects.all().delete()

        for row in rows:
            section_name = row[self.column_index[CARD_SECTION_NAME]]
            if not section_name or section_name == "ei":
                continue

            location = row[self.column_index[CARD_SECTION_LOCATION]]
            show_on_mobile = row[self.column_index[CARD_SHOW_ON_MOBILE]] == "kyllä"
            date_format = row[self.column_index[CARD_SECTION_DATE_FORMAT]]
            attribute = Attribute.objects.get(
                identifier=row[self.column_index[ATTRIBUTE_IDENTIFIER]]
            )

            if not location:
                location = [0]
            elif type(location) == str:
                location = location.split(".")
            elif type(location) == int:
                location = [location]

            attr_index = int(location[-1]) if len(location) >= 1 else 0

            section_keys = {
                "Projektikortin kuva": "projektikortin_kuva",
                "Perustiedot": "perustiedot",
                "Suunnittelualueen kuvaus": "suunnittelualueen_kuvaus",
                "Strategiakytkentä": "strategiakytkenta",
                "Maanomistus ja sopimusmenettelyt": "maanomistus",
                "Kerrosalatiedot": "kerrosalatiedot",
                "Aikataulu": "aikataulu",
                "Yhteyshenkilöt": "yhteyshenkilot",
                "Dokumentit": "dokumentit",
                "Suunnittelualueen rajaus": "suunnittelualueen_rajaus",
            }

            section, __ = ProjectCardSection.objects.update_or_create(
                name=section_name,
                defaults={
                    "index": int(location[0]),
                    "key": section_keys.get(section_name)
                },
            )
            ProjectCardSectionAttribute.objects.create(
                attribute=attribute,
                section=section,
                date_format=date_format or None,
                index=attr_index,
                show_on_mobile=show_on_mobile,
            )

    def _create_overview_filters(self, rows):
        logger.info("\nReplacing project overview filters...")
        OverviewFilter.objects.all().delete()

        for row in rows:
            filter_name = row[self.column_index[OVERVIEW_FILTER_NAME]]
            filters = row[self.column_index[OVERVIEW_FILTERS]] or ""
            filters_by_subtype = OverviewViews.BY_SUBTYPE.value in filters
            filters_on_map = OverviewViews.ON_MAP.value in filters
            filters_floor_area = OverviewViews.FLOOR_AREA.value in filters

            if not filter_name:
                continue

            overview_filter, __ = OverviewFilter.objects.get_or_create(
                name=filter_name,
                defaults={
                    "identifier": self._get_identifier_for_value(filter_name)
                },
            )
            attribute = Attribute.objects.get(
                identifier=row[self.column_index[ATTRIBUTE_IDENTIFIER]]
            )
            OverviewFilterAttribute.objects.create(
                attribute=attribute,
                overview_filter=overview_filter,
                filters_by_subtype=filters_by_subtype,
                filters_floor_area=filters_floor_area,
                filters_on_map=filters_on_map,
            )

    def _create_sections(self, rows, subtype: ProjectSubtype):
        logger.info("\nReplacing sections...")

        ProjectPhaseSection.objects.filter(phase__project_subtype=subtype).delete()

        for phase in ProjectPhase.objects.filter(project_subtype=subtype):

            phase_sections = set()

            for row in rows:
                try:
                    location = self._get_attribute_locations(row, phase.name)
                    section_phase_name = location["label"]
                    section_phase_ingress = location["ingress"]
                    index = location["section_location"]
                except TypeError:
                    continue

                phase_sections.add((section_phase_name, section_phase_ingress, index))

            for phase_section_name, section_phase_ingress, index in phase_sections:
                section = ProjectPhaseSection.objects.create(
                    phase=phase, index=index, name=phase_section_name, ingress=section_phase_ingress
                )
                logger.info(f"Created {section}")

    def _create_floor_area_sections(self, rows, subtype: ProjectSubtype):
        logger.info("\nReplacing floor area sections...")

        ProjectFloorAreaSection.objects \
            .filter(project_subtype=subtype).delete()

        # Get all distinct section names in appearance order
        subtype_sections = []

        for row in rows:
            try:
                section_name = \
                    row[self.column_index[ATTRIBUTE_FLOOR_AREA_SECTION]].strip()
            except AttributeError:
                continue

            if section_name != "ei" and section_name not in subtype_sections:
                subtype_sections.append(section_name)

        for i, section_name in enumerate(subtype_sections, start=1):
            section = ProjectFloorAreaSection.objects.create(
                project_subtype=subtype, index=i, name=section_name
            )
            logger.info(f"Created {section}")

    @staticmethod
    def calculate_deadline_index(locations):
        deadline_index = 0
        for index, location in enumerate(locations, start=1):
            if index == 3:
                deadline_index += int(int(location) * 10000 / int(10 ** 4))
            else:
                deadline_index += int(int(location) * 10000 / int(10 ** index))
        return deadline_index

    @staticmethod
    def calculate_index(locations):
        field_location = locations["field_location"]
        child_locations = locations["child_locations"]
        if child_locations:
            child_index = 0
            for index, child_location in enumerate(child_locations, start=1):
                child_index += int(child_location / int(10 ** index))
            return field_location + child_index
        return field_location

    def _create_attribute_section_links(self, rows, subtype: ProjectSubtype):
        logger.info("\nReplacing attribute section links...")

        ProjectPhaseSectionAttribute.objects.filter(
            section__phase__project_subtype=subtype
        ).delete()

        for phase in ProjectPhase.objects.filter(project_subtype=subtype):

            # Index for attribute within a phase section
            counter = Counter()

            for row in rows:
                project_size = row[self.column_index[PROJECT_SIZE]]
                row_subtypes = self.get_subtypes_from_cell(project_size)

                # Filter out attributes that should not be included in the project sub type (size)
                if (
                    "kaikki" not in row_subtypes
                    and subtype.name.lower() not in row_subtypes
                ):
                    continue

                identifier = self._get_attribute_row_identifier(row)
                attribute = Attribute.objects.get(identifier=identifier)
                locations = self._get_attribute_locations(row, phase.name)

                if locations is None:
                    # Attribute doesn't appear in this phase
                    continue

                section_phase_name = locations["label"]
                section_phase_ingress = locations["ingress"]

                section = ProjectPhaseSection.objects.get(
                    phase=phase,
                    name=section_phase_name,
                    ingress=section_phase_ingress,
                    index=locations["section_location"],
                )

                attribute_index = self.calculate_index(locations)

                section_attribute = ProjectPhaseSectionAttribute.objects.create(
                    attribute=attribute,
                    section=section,
                    index=attribute_index,
                )

                counter[section] += 1

                logger.info(
                    f"Created "
                    f"{section_attribute.section.phase} / "
                    f"{section_attribute.section} / "
                    f"{section_attribute.index} / "
                    f"{section_attribute.attribute}"
                )

    def _create_floor_area_attribute_section_links(
        self, rows, subtype: ProjectSubtype
    ):
        logger.info("\nReplacing attribute floor area section links...")

        ProjectFloorAreaSectionAttribute.objects.filter(
            section__project_subtype=subtype
        ).delete()

        # Index for attribute within a phase section
        counter = Counter()

        for row in rows:
            project_size = row[self.column_index[PROJECT_SIZE]]
            row_subtypes = self.get_subtypes_from_cell(project_size)

            # Filter out attributes that should not be included in the project sub type (size)
            if (
                "kaikki" not in row_subtypes
                and subtype.name.lower() not in row_subtypes
            ):
                continue

            identifier = self._get_attribute_row_identifier(row)
            attribute = Attribute.objects.get(identifier=identifier)

            try:
                section_name = \
                    row[self.column_index[ATTRIBUTE_FLOOR_AREA_SECTION]].strip()
            except AttributeError:
                continue

            if section_name == "ei":
                continue

            section = ProjectFloorAreaSection.objects.get(
                project_subtype=subtype, name=section_name
            )

            section_attribute = ProjectFloorAreaSectionAttribute.objects.create(
                attribute=attribute,
                section=section,
                index=counter[section],
            )

            # Create matrices
            try:
                row_names = \
                    row[self.column_index[ATTRIBUTE_FLOOR_AREA_SECTION_MATRIX_ROW]].strip()

                cell_names = \
                    row[self.column_index[ATTRIBUTE_FLOOR_AREA_SECTION_MATRIX_CELL]].strip()

                if row_names != "ei" and cell_names != "ei":
                    row_names = row_names.split("\n")
                    cell_names = cell_names.split("\n")

                    structure, _ = ProjectFloorAreaSectionAttributeMatrixStructure \
                        .objects.get_or_create(
                            section=section,
                            defaults={
                                'column_names': [],
                                'row_names': [],
                            }
                        )

                    for row_name in row_names:
                        if row_name not in structure.row_names:
                            structure.row_names.append(row_name)
                            structure.save()

                        for cell_name in cell_names:
                            if cell_name not in structure.column_names:
                                structure.column_names.append(cell_name)
                                structure.save()

                            ProjectFloorAreaSectionAttributeMatrixCell.objects.get_or_create(
                                row=structure.row_names.index(row_name),
                                column=structure.column_names.index(cell_name),
                                structure=structure,
                                attribute=section_attribute,
                            )

            except AttributeError:
                pass

            counter[section] += 1

            logger.info(
                f"Created "
                f"{section_attribute.section.project_subtype} / "
                f"{section_attribute.section} / "
                f"{section_attribute.index} / "
                f"{section_attribute.attribute}"
            )

    def _create_deadline_sections(self, rows, subtype):
        logger.info("\nReplacing deadline sections")

        ProjectPhaseDeadlineSection.objects \
            .filter(phase__project_subtype=subtype).delete()

        for row in rows:
            try:
                attribute = Attribute.objects.get(
                    identifier=row[self.column_index[ATTRIBUTE_IDENTIFIER]],
                )
            except Attribute.DoesNotExist:
                continue

            if not attribute.deadline:
                continue

            admin_section = \
                row[self.column_index[ATTRIBUTE_DEADLINE_SECTION_COLUMNS["admin"]]]
            owner_section = \
                row[self.column_index[ATTRIBUTE_DEADLINE_SECTION_COLUMNS["owner"]]]

            for section_string, owner in [(admin_section, False), (owner_section, True)]:
                phase_name_regex = r"(Käynnistys|Periaatteet|OAS|Luonnos|Ehdotus|Tarkistettu ehdotus|Hyväksyminen|Voimaantulo)"

                if not section_string:
                    continue

                try:
                    phase = ProjectPhase.objects.get(
                        common_project_phase__name=re.findall(
                            phase_name_regex, section_string
                        )[0],
                        project_subtype=subtype,
                    )
                except (IndexError, ProjectPhase.DoesNotExist):
                    continue

                section, __ = ProjectPhaseDeadlineSection.objects.get_or_create(
                    phase=phase,
                    defaults={
                        "index": phase.index,
                    }
                )

                if admin_section == owner_section:
                    defaults = {
                        "owner_field": True,
                        "admin_field": True,
                    }
                elif owner:
                    # Marked as owner fields here and in the excel but should actually
                    # be visible (although non-editable) to everyone besides admins
                    # who have their own sections specified
                    defaults = {"owner_field": True}
                else:
                    defaults = {"admin_field": True}

                try:
                    locations = re.split(r";\s*", section_string)[-1].split(".")
                    index = self.calculate_deadline_index(locations)
                except Exception:
                    index = 0

                valid_subtypes = row[self.column_index[PROJECT_SIZE]]
                is_valid_subtype = (
                    valid_subtypes == "kaikki" or
                    not valid_subtypes or
                    subtype.name in re.findall("[A-Z]+", valid_subtypes)
                )
                if is_valid_subtype:
                    ProjectPhaseDeadlineSectionAttribute.objects.get_or_create(
                        attribute=attribute,
                        section=section,
                        defaults=defaults,
                        index=index,
                    )

    def _create_attribute_categorizations(self, rows: Iterable[Sequence[str]]):
        logger.info("\nCreating attribute categorizations")

        attribute_categorizations = {attr.pk: attr for attr in AttributeCategorization.objects.all()}
        common_project_phases = {phase.name: phase for phase in CommonProjectPhase.objects.all()}

        for row in rows:
            attribute_identifier = self._get_attribute_row_identifier(row)
            try:
                attribute = Attribute.objects.get(identifier=attribute_identifier)
                for (column_label, phase, included_phases) in ATTRIBUTE_CATEGORIZATION_COLUMNS:
                    value = row[self.column_index[column_label]]
                    if not value:
                        continue

                    principle = Phases.PRINCIPLES in included_phases
                    draft = Phases.DRAFT in included_phases
                    attribute_categorization, created = AttributeCategorization.objects.update_or_create(
                        attribute=attribute,
                        common_project_phase=common_project_phases.get(phase.value),
                        includes_principles=principle,
                        includes_draft=draft,
                        defaults={"value": value}
                    )
                    try:
                        if not created:
                            del attribute_categorizations[attribute_categorization.pk]
                    except KeyError:
                        pass  # Object doesn't exist in attribute_categorizations dict
            except Attribute.DoesNotExist:
                logger.warning(f"Attribute not found by identifier: {attribute_identifier}")

        for ag in attribute_categorizations.values():
            ag.delete()


    def get_subtypes_from_cell(self, cell_content: Optional[str]) -> List[str]:
        # If the subtype is missing we assume it is to be included in all subtypes
        if not cell_content:
            return ["kaikki"]

        if "," in cell_content:
            # Split names and remove whitespace
            return [name.strip().lower() for name in cell_content.split(",")]
        else:
            # Always handle a list
            return [cell_content.lower()]

    def create_phases(self, subtype: ProjectSubtype):
        logger.info(f"\nCreating phases for {subtype.name}...")
        phase_names = SUBTYPE_PHASES[subtype.name.upper()]

        created_phase_count = 0
        updated_phase_count = 0
        old_phases = [obj.id for obj in subtype.phases.order_by("index")]
        for i, phase_name in enumerate(phase_names, start=1):
            phase = PROJECT_PHASES[phase_name]
            metadata = SUBTYPE_PHASE_METADATA[subtype.name.upper()][phase_name]
            common_phase, _ = CommonProjectPhase.objects.update_or_create(
                name=phase["name"],
                defaults={
                    "index": i,
                    "color": phase["color"],
                    "color_code": phase["color_code"],
                    "list_prefix": phase["list_prefix"],
                }
            )

            if i > common_phase.index:
                common_phase.index = i
                common_phase.save()

            project_phase, created = ProjectPhase.objects.update_or_create(
                project_subtype=subtype,
                common_project_phase=common_phase,
                defaults={
                    "metadata": metadata,
                    "index": common_phase.index,
                },
            )
            if project_phase.id in old_phases:
                old_phases.remove(project_phase.id)

            if created:
                created_phase_count += 1
            else:
                updated_phase_count += 1

        try:
            ProjectPhase.objects.filter(id__in=old_phases).delete()
        except ProtectedError:
            # TODO: Handle removal of Phases when they are in use by Projects
            #       could for instance be disabled by new projects but visible
            #       for old ones.
            pass

        return {
            "created": created_phase_count,
            "updated": updated_phase_count,
            "deleted": len(old_phases),
        }

    def create_subtypes(
        self, rows: Iterable[Sequence[str]]
    ) -> Iterable[ProjectSubtype]:

        subtype_names = [
            subtype.name.lower() for subtype in ProjectSubtype.objects.all()
        ]
        for row in rows:
            subtypes_cell_content = row[self.column_index[PROJECT_SIZE]]
            row_subtypes = self.get_subtypes_from_cell(subtypes_cell_content)

            # If all ("kaikki") of the subtypes are included, continue
            if "kaikki" in row_subtypes:
                continue

            for subtype_name in row_subtypes:
                # Skip non-single word and existing subtypes
                if " " in subtype_name or subtype_name in subtype_names:
                    continue

                subtype_names.append(subtype_name)

        # Sort subtypes by "clothing sizes"
        sort_order = ["xxs", "xs", "s", "m", "l", "xl", "xxl"]
        # Note that if a value is not in the sort order list, it will be ordered first in the list
        ordered_subtype_names = sorted(
            subtype_names,
            key=lambda x: next((i for i, t in enumerate(sort_order) if x == t), 0),
        )

        # Create subtypes
        ordered_subtypes = []
        for index, subtype_name in enumerate(ordered_subtype_names):
            project_subtype, __ = ProjectSubtype.objects.update_or_create(
                pk=index+1,
                project_type=self.project_type,
                name=subtype_name.upper(),
                defaults={"index": index},
            )
            ordered_subtypes.append(project_subtype)

        return ordered_subtypes

    @transaction.atomic
    def run(self):
        self.project_type, _ = ProjectType.objects.get_or_create(name="asemakaava")

        filename = self.options.get("filename")
        logger.info(
            f"Importing attributes from file {filename} for project type {self.project_type}..."
        )

        self.workbook = self._open_workbook(filename)
        rows = self._extract_data_from_workbook(self.workbook)

        header_row = rows[0]
        self._set_row_indexes(header_row)

        data_rows = list(filter(self._check_if_row_valid, rows[1:]))

        subtypes = self.create_subtypes(data_rows)
        attribute_info = self._create_attributes(data_rows)
        self._create_attribute_key_relations(data_rows)
        phase_info = {"created": 0, "updated": 0, "deleted": 0}
        # Reset Fieldset relations
        FieldSetAttribute.objects.all().delete()
        for subtype in subtypes:
            _phase_info = self.create_phases(subtype)
            phase_info["created"] += _phase_info["created"]
            phase_info["updated"] += _phase_info["updated"]
            phase_info["deleted"] += _phase_info["deleted"]
            self._create_fieldset_links(subtype, data_rows)

        self._validate_generated_attributes()

        # Remove all attributes from further processing that were part of a fieldset
        all_data_rows = data_rows
        data_rows = list(filterfalse(self._row_part_of_fieldset, data_rows))

        for subtype in subtypes:
            self._create_sections(data_rows, subtype)
            self._create_attribute_section_links(data_rows, subtype)

            self._create_floor_area_sections(data_rows, subtype)
            self._create_floor_area_attribute_section_links(data_rows, subtype)
            self._create_deadline_sections(data_rows, subtype)

        self._create_card_sections(all_data_rows)
        self._create_document_link_sections(all_data_rows)
        self._create_overview_filters(all_data_rows)
        self._create_attribute_categorizations(all_data_rows)

        # Clear cached sections
        cache.delete("serialized_phase_sections")
        cache.delete("serialized_deadline_sections")

        logger.info("Project subtypes {}".format(ProjectSubtype.objects.count()))
        logger.info("Phases {}".format(ProjectPhase.objects.count()))
        logger.info(f"  Created: {phase_info['created']}")
        logger.info(f"  Updated: {phase_info['updated']}")
        logger.info(f"  Deleted: {phase_info['deleted']}")
        logger.info("Attributes {}".format(Attribute.objects.count()))
        logger.info(f"  Created: {attribute_info['created']}")
        logger.info(f"  Updated: {attribute_info['updated']}")
        logger.info(f"  Deleted: {attribute_info['deleted']}")
        logger.info(f"  Choices: {attribute_info['choices']}")
        logger.info(
            "FieldSets {}".format(
                Attribute.objects.filter(value_type=Attribute.TYPE_FIELDSET).count()
            )
        )
        logger.info(
            "Info FieldSets {}".format(
                Attribute.objects.filter(value_type=Attribute.TYPE_INFO_FIELDSET).count()
            )
        )
        logger.info("Phase sections {}".format(ProjectPhaseSection.objects.count()))
        logger.info(
            "Phase section attributes {}".format(
                ProjectPhaseSectionAttribute.objects.count()
            )
        )
        logger.info("Import done.")
